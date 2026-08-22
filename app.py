from __future__ import annotations

import copy
import csv
import hashlib
import hmac
import json
import os
import re
import secrets
import shutil
import tempfile
import threading
import time
from collections import OrderedDict
from contextlib import contextmanager
from urllib import error as urlerror
from urllib import parse as urlparse
from urllib import request as urlrequest
from functools import wraps
from datetime import timedelta
from io import BytesIO, StringIO
from uuid import uuid4

from flask import (Flask, abort, g, jsonify, make_response, redirect, render_template,
                   request, send_file, session, url_for)
from PIL import Image, ImageDraw, ImageFont
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import BadSignature, URLSafeSerializer
from dotenv import load_dotenv
import yaml

import db

load_dotenv()
app = Flask(__name__)

# Certificate links are signed with this key, so an unstable key does not just log
# admins out at random - it invalidates every outstanding certificate link too.
_SECRET_KEY = os.environ.get("SECRET_KEY", "").strip()
if not _SECRET_KEY:
	_SECRET_KEY = os.urandom(24).hex()
	print(
		"WARNING: SECRET_KEY is not set. A random key was generated for this process, "
		"so sessions and certificate links will break across restarts and will not "
		"work at all with more than one worker. Set SECRET_KEY in the environment."
	)
app.secret_key = _SECRET_KEY

# The club/event store. make_repository() picks Postgres when DATABASE_URL is set
# and an in-memory store otherwise, announcing which at startup. Held in a module
# global so tests can inject a backend (a real in-memory one, or a stub that
# raises to exercise the Postgres-down path).
_repository = db.make_repository()


def repo() -> "db.Repository":
	return _repository


@app.errorhandler(db.DatabaseUnavailable)
def _database_unavailable(_exc):
	"""Postgres is configured but unreachable. Fail closed - never serve stale or
	empty account data - with a 503 the client can retry."""
	if request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.is_json:
		response = jsonify({"ok": False, "error": "The service is temporarily unavailable."})
	else:
		response = make_response(
			render_template("service_unavailable.html")
			if os.path.exists(os.path.join(BASE_DIR, "templates", "service_unavailable.html"))
			else "Service temporarily unavailable. Please try again shortly.", 503)
	response.status_code = 503
	response.headers["Retry-After"] = "10"
	response.cache_control.no_store = True
	return response
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB upload limit
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "").strip().lower() in ("1", "true", "yes")
# Logins persist across browser restarts: sessions marked permanent get a signed
# cookie with this rolling lifetime (refreshed each request) instead of dying when
# the browser closes. Auth stays in the HTTPOnly cookie - never in localStorage.
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_EVENTS_DIR = os.path.join(BASE_DIR, "events")


def _resolve_runtime_writable_dir() -> str:
	"""Use BASE_DIR when writable, otherwise fall back to OS temp directory."""
	probe_path = os.path.join(BASE_DIR, ".write_probe")
	try:
		with open(probe_path, "w", encoding="utf-8") as f:
			f.write("ok")
		os.remove(probe_path)
		return BASE_DIR
	except OSError:
		return tempfile.gettempdir()


# Serverless deployments may have a read-only code directory.
RUNTIME_WRITABLE_DIR = _resolve_runtime_writable_dir()
EVENTS_DIR = os.path.join(RUNTIME_WRITABLE_DIR, "events")
GENERATED_DIR = os.path.join(RUNTIME_WRITABLE_DIR, "generated_certificates")
# Background jobs log here. It lives on the writable runtime dir because the code
# directory is read-only on some hosts.
LOG_DIR = os.path.join(RUNTIME_WRITABLE_DIR, "logs")
LOG_FILE = os.path.join(LOG_DIR, "system.log")


def log_event(category: str, message: str) -> None:
	"""Append a line to the admin-visible system log. Best-effort: a read-only or
	full filesystem must never turn a log write into a request failure."""
	import datetime
	try:
		os.makedirs(LOG_DIR, exist_ok=True)
		with open(LOG_FILE, "a", encoding="utf-8") as f:
			stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
			f.write(f"[{stamp}] [{category}] {message}\n")
	except OSError:
		pass
FONT_PATH = os.path.join(BASE_DIR, "fonts", "Montserrat-Bold.ttf")
DEFAULT_FONT_KEY = "montserrat_bold"
FONT_OPTIONS = {
	"montserrat_bold": {
		"label": "Montserrat Bold",
		"filename": "Montserrat-Bold.ttf",
		"css_family": "Montserrat Bold",
		"css_weight": "700",
	}
}

EVENT_STATE_FILE = os.path.join(GENERATED_DIR, "event_states.json")
KV_REST_API_URL = os.environ.get("KV_REST_API_URL", "").strip()
KV_REST_API_TOKEN = os.environ.get("KV_REST_API_TOKEN", "").strip()
KV_EVENT_STATE_KEY = os.environ.get("KV_EVENT_STATE_KEY", "certificate_generator:event_states")
KV_EVENT_INDEX_KEY = os.environ.get("KV_EVENT_INDEX_KEY", "certificate_generator:event_index")
KV_EVENT_CONFIG_PREFIX = os.environ.get("KV_EVENT_CONFIG_PREFIX", "certificate_generator:event_config:")
KV_EVENT_CSV_PREFIX = os.environ.get("KV_EVENT_CSV_PREFIX", "certificate_generator:event_csv:")
_EVENT_STATE_CACHE: dict[str, dict] | None = None
_EVENT_STATE_CACHE_AT = 0.0
_EVENT_STATE_CACHE_TTL_SEC = 2.0

# Supabase Storage holds certificate templates so they survive redeploys and are
# visible to every worker. Local files stay as the dev / no-Supabase fallback.
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").strip().rstrip("/")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "").strip()
# One bucket for the club, one folder per event inside it. Supabase renders key
# prefixes as folders, so the dashboard shows <event>/participants and
# <event>/template. Bucket names cannot contain spaces, hence the slug form.
SUPABASE_BUCKET = os.environ.get("SUPABASE_BUCKET", "csi-aseb").strip()
_SUPABASE_BUCKET_READY = False


def _env_int(name: str, fallback: int) -> int:
	try:
		return max(1, int(os.environ.get(name, "").strip()))
	except (TypeError, ValueError):
		return fallback


# Event config cache (per-slug) to avoid repeated KV API calls during preview/download
_EVENT_CONFIG_CACHE: dict[str, tuple[dict, float]] = {}  # {slug: (config, timestamp)}
_EVENT_CONFIG_CACHE_TTL_SEC = 30.0  # 30 second TTL for event configs

# Participant CSVs are read several times per page (team list, each dropdown column,
# then again during validation). Without this, each of those was a separate KV
# round trip. Strings are immutable, so the cached value can be shared freely.
_EVENT_CSV_CACHE: dict[str, tuple[str | None, float]] = {}  # {slug: (csv_text, timestamp)}
_EVENT_CSV_CACHE_TTL_SEC = 30.0

# A decoded template is tens of megabytes (an A4 300 DPI page is 8.7 MP), so these
# caches stay small. Templates are held as RGB unless they really carry alpha,
# which is 25% less memory than RGBA and buys a couple more cache slots.
#
# The template cache is bounded by bytes, not entries: six A4 300 DPI pages is
# 150 MB while six phone-sized ones is 3 MB, so an entry count is a guess at the
# thing that actually runs out. With many clubs cycling many distinct templates
# through one cache, that guess is what OOM-kills the worker.
TEMPLATE_CACHE_MAX_BYTES = _env_int("TEMPLATE_CACHE_MAX_BYTES", 96 * 1024 * 1024)
_TEMPLATE_CACHE_MAX = _env_int("TEMPLATE_CACHE_MAX", 6)
_RENDER_CACHE_MAX = _env_int("RENDER_CACHE_MAX", 6)
_FONT_CACHE_MAX = 20

# Encoding dominates render time: for an 8.7 MP page, drawing the name takes ~4 ms
# while the PNG encode takes ~320 ms at Pillow's default compression. So the screen
# preview is downscaled and sent as JPEG, and the download drops to a lighter PNG
# compression level - which, with the alpha channel gone, is still a smaller file.
PREVIEW_MAX_WIDTH = _env_int("PREVIEW_MAX_WIDTH", 1200)
PREVIEW_JPEG_QUALITY = _env_int("PREVIEW_JPEG_QUALITY", 90)
DOWNLOAD_PNG_COMPRESS_LEVEL = _env_int("DOWNLOAD_PNG_COMPRESS_LEVEL", 3)

# Download format, per event. PNG at any compression level costs 250-530 ms for a
# 3508x2480 page; JPEG q92 costs ~45 ms and produces 3.5x fewer bytes, which is the
# difference between needing a core and needing a tenth of one at 1.7 arrivals/sec.
#
# An absent "download_format" means PNG, so every event that predates this change
# renders byte-identically with no migration. New events are created with "jpeg"
# written explicitly, so "absent" only ever describes a legacy config.
DOWNLOAD_FORMATS = ("png", "jpeg")
DEFAULT_DOWNLOAD_FORMAT = "png"
NEW_EVENT_DOWNLOAD_FORMAT = "jpeg"
DOWNLOAD_JPEG_QUALITY = _env_int("DOWNLOAD_JPEG_QUALITY", 92)
# 4:4:4. JPEG's default 4:2:0 halves chroma resolution and bleeds colour around
# sharp text edges: measured RMS error over the text box roughly doubles for
# saturated text (1.68 -> 3.51), which is exactly what a club's brand colour is.
# Costs ~10 ms and ~200 KB per render, leaving it 6x cheaper than PNG.
DOWNLOAD_JPEG_SUBSAMPLING = 0
DOWNLOAD_MIMETYPE_EXTENSIONS = {"image/png": ".png", "image/jpeg": ".jpg"}

class _ByteCappedCache:
	"""
	LRU cache bounded by total stored bytes as well as by entry count.

	Entry count is the wrong bound for anything whose entries differ by orders of
	magnitude in size. Bytes make the ceiling a RAM budget instead of a guess; the
	entry cap is kept as a secondary bound so an existing TEMPLATE_CACHE_MAX in a
	deployment's environment still means what it used to.
	"""

	def __init__(self, max_bytes: int, max_entries: int):
		self.max_bytes = max_bytes
		self.max_entries = max_entries
		self.total_bytes = 0
		self._entries: "OrderedDict[str, tuple]" = OrderedDict()

	def get(self, key: str):
		entry = self._entries.get(key)
		if entry is None:
			return None
		self._entries.move_to_end(key)
		return entry[0]

	def put(self, key: str, value, nbytes: int) -> None:
		previous = self._entries.pop(key, None)
		if previous is not None:
			self.total_bytes -= previous[1]
		self._entries[key] = (value, nbytes)
		self.total_bytes += nbytes
		# An entry larger than the whole budget is evicted immediately rather than
		# held: better to re-decode it than to blow the budget it was sized against.
		while self._entries and (self.total_bytes > self.max_bytes
								 or len(self._entries) > self.max_entries):
			_, evicted = self._entries.popitem(last=False)
			self.total_bytes -= evicted[1]

	def clear(self) -> None:
		self._entries.clear()
		self.total_bytes = 0

	def __len__(self) -> int:
		return len(self._entries)

	def __contains__(self, key: str) -> bool:
		return key in self._entries


# Every cache key embeds the template version and render settings, so a replaced
# template or an edited coordinate produces a new key instead of a stale hit.
_TEMPLATE_IMAGE_CACHE = _ByteCappedCache(TEMPLATE_CACHE_MAX_BYTES, _TEMPLATE_CACHE_MAX)
_RENDERED_CERT_CACHE: "OrderedDict[str, tuple[bytes, str]]" = OrderedDict()
_FONT_CACHE: "OrderedDict[str, ImageFont.FreeTypeFont | ImageFont.ImageFont]" = OrderedDict()


def _cache_get(cache: OrderedDict, key: str):
	if key not in cache:
		return None
	cache.move_to_end(key)
	return cache[key]


def _cache_put(cache: OrderedDict, key: str, value, max_size: int) -> None:
	cache[key] = value
	cache.move_to_end(key)
	while len(cache) > max_size:
		cache.popitem(last=False)


# ─── Render concurrency ───────────────────────────────────────────────────────
#
# A render is CPU-bound Pillow work holding a ~25 MB image copy for its duration,
# so unbounded concurrency does not serve a spike faster - it OOM-kills the worker
# and everyone loses. Requests queue here instead. The per-tenant limit stops one
# busy event from taking every slot on the instance.
#
# Tenants are events today; in Phase 1 the key becomes the club id, which is why
# render_certificate takes a tenant_key instead of assuming the slug.

def _cgroup_cpu_quota() -> float | None:
	"""CPU quota this container is allowed, in cores, or None if unrestricted."""
	try:  # cgroup v2
		with open("/sys/fs/cgroup/cpu.max", encoding="utf-8") as f:
			quota, period = f.read().split()
		if quota != "max" and float(period) > 0:
			return float(quota) / float(period)
	except (OSError, ValueError):
		pass
	try:  # cgroup v1
		with open("/sys/fs/cgroup/cpu/cpu.cfs_quota_us", encoding="utf-8") as f:
			quota = int(f.read().strip())
		with open("/sys/fs/cgroup/cpu/cpu.cfs_period_us", encoding="utf-8") as f:
			period = int(f.read().strip())
		if quota > 0 and period > 0:
			return quota / period
	except (OSError, ValueError):
		pass
	return None


def available_cores() -> int:
	"""
	Cores this process may actually use.

	os.cpu_count() reports the *host's* cores, not the container's quota, so on a
	fractional-CPU instance it over-reports by an order of magnitude - which is
	precisely the over-subscription the semaphore exists to prevent.
	"""
	quota = _cgroup_cpu_quota()
	if quota is not None:
		return max(1, int(quota))
	return max(1, os.cpu_count() or 1)


RENDER_MAX_CONCURRENCY = _env_int("RENDER_MAX_CONCURRENCY", available_cores())
RENDER_MAX_CONCURRENCY_PER_TENANT = _env_int(
	"RENDER_MAX_CONCURRENCY_PER_TENANT", max(1, RENDER_MAX_CONCURRENCY // 2))
# Bounded so a spike sheds load with a 503 instead of piling up until gunicorn's
# --timeout kills the worker and takes every in-flight request with it.
RENDER_QUEUE_TIMEOUT_SEC = _env_int("RENDER_QUEUE_TIMEOUT_SEC", 30)

_RENDER_SLOTS = threading.BoundedSemaphore(RENDER_MAX_CONCURRENCY)
_TENANT_SLOTS: dict[str, threading.BoundedSemaphore] = {}
_TENANT_SLOTS_LOCK = threading.Lock()


class RenderCapacityError(RuntimeError):
	"""No render slot came free before the queue timeout. Serve a 503, not a 404."""


def configure_render_slots(max_concurrency: int | None = None,
						   per_tenant: int | None = None) -> None:
	"""Resize the render semaphores. Used by the load test and the test suite."""
	global RENDER_MAX_CONCURRENCY, RENDER_MAX_CONCURRENCY_PER_TENANT, _RENDER_SLOTS
	if max_concurrency is not None:
		RENDER_MAX_CONCURRENCY = max(1, max_concurrency)
	if per_tenant is not None:
		RENDER_MAX_CONCURRENCY_PER_TENANT = max(1, per_tenant)
	_RENDER_SLOTS = threading.BoundedSemaphore(RENDER_MAX_CONCURRENCY)
	with _TENANT_SLOTS_LOCK:
		_TENANT_SLOTS.clear()


def _tenant_slots(tenant_key: str) -> threading.BoundedSemaphore:
	with _TENANT_SLOTS_LOCK:
		slots = _TENANT_SLOTS.get(tenant_key)
		if slots is None:
			slots = threading.BoundedSemaphore(RENDER_MAX_CONCURRENCY_PER_TENANT)
			_TENANT_SLOTS[tenant_key] = slots
		return slots


@contextmanager
def render_slot(tenant_key: str):
	"""
	Hold one per-tenant and one global render slot, or raise RenderCapacityError.

	The tenant slot is taken first: a tenant already at its own limit would
	otherwise sit on a scarce global slot while it waited for its own. The
	acquisition order is fixed, so the two semaphores cannot deadlock.
	"""
	timeout = max(0, RENDER_QUEUE_TIMEOUT_SEC)
	deadline = time.monotonic() + timeout
	tenant = _tenant_slots(tenant_key)
	if not tenant.acquire(timeout=timeout):
		raise RenderCapacityError(f"no render slot free for '{tenant_key}'")
	try:
		if not _RENDER_SLOTS.acquire(timeout=max(0, deadline - time.monotonic())):
			raise RenderCapacityError("render capacity exhausted")
		try:
			yield
		finally:
			_RENDER_SLOTS.release()
	finally:
		tenant.release()

os.makedirs(EVENTS_DIR, exist_ok=True)
os.makedirs(GENERATED_DIR, exist_ok=True)

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")

# Certificate tokens minted before Phase 3 carry no club (`c`). They resolve to
# this club during a grace period; every such resolution is logged so the log
# going quiet is the signal the grace period can end. Set LEGACY_TOKEN_GRACE=0 to
# reject club-less tokens outright (the cutoff).
LEGACY_TOKEN_CLUB = "csi-aseb"
LEGACY_TOKEN_GRACE = os.environ.get("LEGACY_TOKEN_GRACE", "1").strip().lower() not in ("0", "false", "no", "off")
TEMPLATE_EXTENSIONS = (".png", ".jpg", ".jpeg", ".gif", ".webp")
TEMPLATE_CONTENT_TYPES = {
	".png": "image/png",
	".jpg": "image/jpeg",
	".jpeg": "image/jpeg",
	".gif": "image/gif",
	".webp": "image/webp",
}
# Guards against decompression bombs: a 40 MP RGBA image is already ~160 MB decoded.
MAX_TEMPLATE_PIXELS = 40_000_000

PARTICIPANT_EXTENSIONS = (".csv", ".xlsx")
PARTICIPANT_CONTENT_TYPES = {
	".csv": "text/csv",
	".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
_SLUG_RE = re.compile(r"^[a-z0-9][a-z0-9-]*$")
VALIDATION_TYPES = {"player_team", "name_only", "email", "badge_id", "custom", "none"}
VALIDATION_TYPE_LABELS = {
	"player_team": "Player + Team",
	"name_only": "Name Only",
	"email": "Email",
	"badge_id": "Roll No",
	"custom": "Custom Fields",
	"none": "No Validation",
}


def _kv_enabled() -> bool:
	return bool(
		KV_REST_API_TOKEN
		and KV_REST_API_URL
		and KV_REST_API_URL.lower().startswith(("http://", "https://"))
	)


def _read_event_states_from_file() -> dict[str, dict]:
	if not os.path.exists(EVENT_STATE_FILE):
		return {}
	try:
		with open(EVENT_STATE_FILE, encoding="utf-8") as f:
			loaded = json.load(f)
		if isinstance(loaded, dict):
			return {str(k): v for k, v in loaded.items() if isinstance(v, dict)}
	except Exception:
		return {}
	return {}


def _write_event_states_to_file(states: dict[str, dict]) -> None:
	os.makedirs(os.path.dirname(EVENT_STATE_FILE), exist_ok=True)
	with open(EVENT_STATE_FILE, "w", encoding="utf-8") as f:
		json.dump(states, f, indent=2)


def _kv_get_event_states() -> dict[str, dict]:
	url = f"{KV_REST_API_URL.rstrip('/')}/get/{urlparse.quote(KV_EVENT_STATE_KEY, safe='')}"
	req = urlrequest.Request(url, headers={"Authorization": f"Bearer {KV_REST_API_TOKEN}"})
	with urlrequest.urlopen(req, timeout=5) as response:
		payload = json.loads(response.read().decode("utf-8"))
	raw = payload.get("result")
	if raw in (None, ""):
		return {}
	if isinstance(raw, str):
		loaded = json.loads(raw)
	elif isinstance(raw, dict):
		loaded = raw
	else:
		return {}
	if not isinstance(loaded, dict):
		return {}
	return {str(k): v for k, v in loaded.items() if isinstance(v, dict)}


def _kv_set_event_states(states: dict[str, dict]) -> None:
	encoded_states = json.dumps(states, separators=(",", ":"))
	key = urlparse.quote(KV_EVENT_STATE_KEY, safe="")
	value = urlparse.quote(encoded_states, safe="")
	url = f"{KV_REST_API_URL.rstrip('/')}/set/{key}/{value}"
	req = urlrequest.Request(url, method="POST", headers={"Authorization": f"Bearer {KV_REST_API_TOKEN}"})
	with urlrequest.urlopen(req, timeout=5):
		return


def _kv_get_raw(key: str):
	url = f"{KV_REST_API_URL.rstrip('/')}/get/{urlparse.quote(key, safe='')}"
	req = urlrequest.Request(url, headers={"Authorization": f"Bearer {KV_REST_API_TOKEN}"})
	with urlrequest.urlopen(req, timeout=5) as response:
		payload = json.loads(response.read().decode("utf-8"))
	return payload.get("result")


def _kv_set_raw(key: str, value: str) -> None:
	encoded_value = urlparse.quote(value, safe="")
	url = f"{KV_REST_API_URL.rstrip('/')}/set/{urlparse.quote(key, safe='')}/{encoded_value}"
	req = urlrequest.Request(url, method="POST", headers={"Authorization": f"Bearer {KV_REST_API_TOKEN}"})
	with urlrequest.urlopen(req, timeout=5):
		return


def _kv_delete_key(key: str) -> None:
	url = f"{KV_REST_API_URL.rstrip('/')}/del/{urlparse.quote(key, safe='')}"
	req = urlrequest.Request(url, method="POST", headers={"Authorization": f"Bearer {KV_REST_API_TOKEN}"})
	with urlrequest.urlopen(req, timeout=5):
		return


# ─── Supabase Storage ─────────────────────────────────────────────────────────

def _supabase_enabled() -> bool:
	return bool(
		SUPABASE_URL
		and SUPABASE_SERVICE_KEY
		and SUPABASE_URL.lower().startswith(("http://", "https://"))
	)


def _supabase_headers(extra: dict[str, str] | None = None) -> dict[str, str]:
	headers = {
		"Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
		"apikey": SUPABASE_SERVICE_KEY,
	}
	if extra:
		headers.update(extra)
	return headers


def _supabase_object_url(object_path: str) -> str:
	bucket = urlparse.quote(SUPABASE_BUCKET, safe="")
	encoded_path = urlparse.quote(object_path, safe="/")
	return f"{SUPABASE_URL}/storage/v1/object/{bucket}/{encoded_path}"


def _supabase_ensure_bucket() -> None:
	"""Create the private storage bucket on first use. Safe to call repeatedly."""
	global _SUPABASE_BUCKET_READY
	if _SUPABASE_BUCKET_READY:
		return
	payload = json.dumps({"id": SUPABASE_BUCKET, "name": SUPABASE_BUCKET, "public": False}).encode("utf-8")
	req = urlrequest.Request(
		f"{SUPABASE_URL}/storage/v1/bucket",
		data=payload,
		method="POST",
		headers=_supabase_headers({"Content-Type": "application/json"}),
	)
	try:
		with urlrequest.urlopen(req, timeout=10):
			pass
	except urlerror.HTTPError as exc:
		# 400/409 is the steady state: the bucket already exists.
		if exc.code not in (400, 409):
			raise
	_SUPABASE_BUCKET_READY = True


def _supabase_ping() -> None:
	"""Round-trip to Supabase so a keep-alive cron also keeps the project awake."""
	bucket = urlparse.quote(SUPABASE_BUCKET, safe="")
	req = urlrequest.Request(f"{SUPABASE_URL}/storage/v1/bucket/{bucket}", headers=_supabase_headers())
	with urlrequest.urlopen(req, timeout=10):
		return


def _supabase_upload(object_path: str, data: bytes, content_type: str) -> None:
	_supabase_ensure_bucket()
	req = urlrequest.Request(
		_supabase_object_url(object_path),
		data=data,
		method="POST",
		headers=_supabase_headers({"Content-Type": content_type, "x-upsert": "true"}),
	)
	with urlrequest.urlopen(req, timeout=30):
		return


def _supabase_download(object_path: str) -> bytes | None:
	req = urlrequest.Request(_supabase_object_url(object_path), headers=_supabase_headers())
	try:
		with urlrequest.urlopen(req, timeout=20) as response:
			return response.read()
	except urlerror.HTTPError as exc:
		if exc.code == 404:
			return None
		raise


def _supabase_delete(object_path: str) -> None:
	req = urlrequest.Request(_supabase_object_url(object_path), method="DELETE", headers=_supabase_headers())
	try:
		with urlrequest.urlopen(req, timeout=15):
			return
	except urlerror.HTTPError as exc:
		if exc.code != 404:
			raise


def _load_event_states(force: bool = False) -> dict[str, dict]:
	global _EVENT_STATE_CACHE, _EVENT_STATE_CACHE_AT
	if not force and _EVENT_STATE_CACHE is not None and (time.time() - _EVENT_STATE_CACHE_AT) < _EVENT_STATE_CACHE_TTL_SEC:
		return _EVENT_STATE_CACHE
	states: dict[str, dict]
	if _kv_enabled():
		try:
			states = _kv_get_event_states()
		except (urlerror.URLError, TimeoutError, json.JSONDecodeError, OSError, ValueError):
			states = _read_event_states_from_file()
	else:
		states = _read_event_states_from_file()
	_EVENT_STATE_CACHE = states
	_EVENT_STATE_CACHE_AT = time.time()
	return states


def _save_event_states(states: dict[str, dict]) -> None:
	global _EVENT_STATE_CACHE, _EVENT_STATE_CACHE_AT
	if _kv_enabled():
		try:
			_kv_set_event_states(states)
		except (urlerror.URLError, TimeoutError, OSError, ValueError):
			# Keep local fallback for local dev / temporary outages.
			_write_event_states_to_file(states)
	else:
		_write_event_states_to_file(states)
	_EVENT_STATE_CACHE = states
	_EVENT_STATE_CACHE_AT = time.time()


def _event_state(slug: str, states: dict[str, dict] | None = None) -> dict:
	if states is None:
		states = _load_event_states()
	state = states.get(slug)
	return state if isinstance(state, dict) else {}


def _set_event_state(slug: str, **updates) -> None:
	states = _load_event_states()
	current = _event_state(slug, states)
	next_state = dict(current)
	next_state.update(updates)
	states[slug] = next_state
	_save_event_states(states)
	_ensure_event_slug_registered(slug)


def _bootstrap_runtime_events() -> None:
	"""Copy bundled events into the runtime-writable directory when needed."""
	if EVENTS_DIR == SOURCE_EVENTS_DIR or not os.path.isdir(SOURCE_EVENTS_DIR):
		return
	states = _load_event_states()
	for slug in os.listdir(SOURCE_EVENTS_DIR):
		source_path = os.path.join(SOURCE_EVENTS_DIR, slug)
		target_path = os.path.join(EVENTS_DIR, slug)
		if not os.path.isdir(source_path):
			continue
		if _event_state(slug, states).get("deleted", False):
			continue
		if os.path.exists(os.path.join(target_path, "config.json")):
			continue
		shutil.copytree(source_path, target_path, dirs_exist_ok=True)


# ─── First-run migration ──────────────────────────────────────────────────────

def _migrate_legacy_event() -> None:
	"""Copy certificate_template.png + data.csv into events/think-run-debug/ on first run."""
	slug = "think-run-debug"
	edir = os.path.join(EVENTS_DIR, slug)
	config_path = os.path.join(edir, "config.json")
	if os.path.exists(config_path):
		return
	os.makedirs(edir, exist_ok=True)
	legacy_template = os.path.join(BASE_DIR, "certificate_template.png")
	if os.path.exists(legacy_template):
		shutil.copy2(legacy_template, os.path.join(edir, "template.png"))
	legacy_csv = os.path.join(BASE_DIR, "data.csv")
	if os.path.exists(legacy_csv):
		shutil.copy2(legacy_csv, os.path.join(edir, "data.csv"))
	config = {
		"name": "Think, Run, Debug",
		"slug": slug,
		"active": True,
		"validation_type": "player_team",
		"text_x": 1789,
		"text_y": 1440,
		"font_size": 100,
		"font_color": [50, 34, 24],
		"font_key": DEFAULT_FONT_KEY,
	}
	with open(config_path, "w", encoding="utf-8") as f:
		json.dump(config, f, indent=2)

_bootstrap_runtime_events()
_migrate_legacy_event()


# ─── Event helpers ────────────────────────────────────────────────────────────

def safe_slug(slug: str) -> bool:
	return bool(_SLUG_RE.match(slug)) and ".." not in slug and len(slug) <= 80


def _event_dir(slug: str) -> str:
	return os.path.join(EVENTS_DIR, slug)


def _event_config_path(slug: str) -> str:
	return os.path.join(_event_dir(slug), "config.json")


def _event_config_key(slug: str) -> str:
	return f"{KV_EVENT_CONFIG_PREFIX}{slug}"


def _event_template_path(slug: str) -> str:
	"""Local template path. Returns the first matching image file in the event directory."""
	event_dir = _event_dir(slug)
	for ext in TEMPLATE_EXTENSIONS:
		path = os.path.join(event_dir, f"template{ext}")
		if os.path.exists(path):
			return path
	return os.path.join(event_dir, "template.png")


def _event_csv_path(slug: str) -> str:
	return os.path.join(_event_dir(slug), "data.csv")


def _event_csv_key(slug: str) -> str:
	return f"{KV_EVENT_CSV_PREFIX}{slug}"


# ─── Certificate template storage ─────────────────────────────────────────────

def _storage_event_prefix(event_slug: str, club_slug: str | None = None) -> str:
	"""The object-store prefix for an event: <club>/<event> for a club-owned event,
	bare <event> for a legacy (not-yet-migrated) one."""
	return f"{club_slug}/{event_slug}" if club_slug else event_slug


def _template_object_path(slug: str, ext: str, club_slug: str | None = None) -> str:
	return f"{_storage_event_prefix(slug, club_slug)}/template/template{ext}"


def _legacy_template_object_path(slug: str, ext: str) -> str:
	"""Where templates lived before the per-event folder layout (legacy events only)."""
	return f"events/{slug}/template{ext}"


def _participants_object_path(slug: str, filename: str = "data.csv",
							  club_slug: str | None = None) -> str:
	return f"{_storage_event_prefix(slug, club_slug)}/participants/{filename}"


def _scoped_event_dir(event_slug: str, club_slug: str | None = None) -> str:
	"""Local fallback directory, mirroring the object-store prefix."""
	return os.path.join(EVENTS_DIR, club_slug, event_slug) if club_slug else _event_dir(event_slug)


def template_ext_for(slug: str, config: dict | None = None) -> str:
	"""Extension recorded on the config, falling back to whatever is on local disk."""
	recorded = (config or {}).get("template_ext")
	if isinstance(recorded, str) and recorded.lower() in TEMPLATE_EXTENSIONS:
		return recorded.lower()
	for ext in TEMPLATE_EXTENSIONS:
		if os.path.exists(os.path.join(_event_dir(slug), f"template{ext}")):
			return ext
	return ".png"


def template_version_for(slug: str, config: dict | None = None) -> str:
	"""Cache-busting token for a template. Changes whenever the template changes."""
	version = (config or {}).get("template_version")
	if isinstance(version, str) and version:
		return version
	# Events created before versioning fall back to the local file mtime.
	try:
		return str(int(os.path.getmtime(_event_template_path(slug))))
	except OSError:
		return "none"


def has_template(slug: str, config: dict | None = None) -> bool:
	if (config or {}).get("template_version"):
		return True
	return os.path.exists(_event_template_path(slug))


def load_template_bytes(slug: str, config: dict | None = None,
						club_slug: str | None = None) -> bytes | None:
	ext = template_ext_for(slug, config)
	if _supabase_enabled():
		# A club event lives only in the club scheme, so there is no legacy path to
		# fall back to; a legacy event keeps its new-then-older fallback.
		object_paths = ((_template_object_path(slug, ext, club_slug),) if club_slug
						else (_template_object_path(slug, ext), _legacy_template_object_path(slug, ext)))
		for object_path in object_paths:
			try:
				data = _supabase_download(object_path)
				if data:
					return data
			except (urlerror.URLError, TimeoutError, OSError, ValueError):
				break
	path = os.path.join(_scoped_event_dir(slug, club_slug), f"template{ext}")
	if not os.path.exists(path) and not club_slug:
		path = _event_template_path(slug)
	try:
		with open(path, "rb") as f:
			return f.read()
	except OSError:
		return None


def save_template_bytes(slug: str, data: bytes, ext: str,
						club_slug: str | None = None) -> str:
	"""Persist a template everywhere and return its new version token."""
	if _supabase_enabled():
		_supabase_upload(
			_template_object_path(slug, ext, club_slug),
			data,
			TEMPLATE_CONTENT_TYPES.get(ext, "application/octet-stream"),
		)
	# Keep a local copy too: it is the fallback when Supabase is not configured.
	local_dir = _scoped_event_dir(slug, club_slug)
	try:
		os.makedirs(local_dir, exist_ok=True)
		for stale_ext in TEMPLATE_EXTENSIONS:
			if stale_ext == ext:
				continue
			stale_path = os.path.join(local_dir, f"template{stale_ext}")
			if os.path.exists(stale_path):
				os.remove(stale_path)
		with open(os.path.join(local_dir, f"template{ext}"), "wb") as f:
			f.write(data)
	except OSError:
		# A read-only filesystem is fine as long as Supabase accepted the upload.
		if not _supabase_enabled():
			raise
	return hashlib.sha256(data).hexdigest()[:16]


def delete_event_objects(slug: str, club_slug: str | None = None) -> None:
	"""Remove an event's whole folder: template, participant files, and legacy keys."""
	if not _supabase_enabled():
		return
	paths = [_participants_object_path(slug, club_slug=club_slug)]
	for ext in TEMPLATE_EXTENSIONS:
		paths.append(_template_object_path(slug, ext, club_slug))
		if not club_slug:
			paths.append(_legacy_template_object_path(slug, ext))
	for ext in PARTICIPANT_EXTENSIONS:
		paths.append(_participants_object_path(slug, f"source{ext}", club_slug=club_slug))
	for object_path in paths:
		try:
			_supabase_delete(object_path)
		except (urlerror.URLError, TimeoutError, OSError, ValueError):
			continue


def _supabase_list(prefix: str) -> list[dict]:
	"""One level of objects under a prefix. Files carry metadata (with size);
	folders come back with null metadata."""
	body = json.dumps({"prefix": prefix, "limit": 1000, "offset": 0,
					   "sortBy": {"column": "name", "order": "asc"}}).encode("utf-8")
	url = f"{SUPABASE_URL}/storage/v1/object/list/{urlparse.quote(SUPABASE_BUCKET, safe='')}"
	req = urlrequest.Request(url, data=body, method="POST",
							 headers=_supabase_headers({"Content-Type": "application/json"}))
	with urlrequest.urlopen(req, timeout=20) as response:
		payload = json.loads(response.read())
	return payload if isinstance(payload, list) else []


def club_storage_bytes(club_slug: str, strict: bool = False) -> int:
	"""Total bytes stored under a club's prefix.

	The object listing is ground truth. It is summed on demand (uploads are rare
	admin actions, not a hot path), walking the club's event/template/participant
	folders. When Supabase is not configured, the local fallback tree is summed.

	`strict=True` re-raises a listing failure instead of silently dropping that
	subtree's bytes. The quota check uses it so a Storage outage fails closed
	(reject the upload) rather than under-counting and admitting it; the dashboard
	display uses the lenient default so a transient error shows a partial figure
	instead of an error page.
	"""
	if not club_slug:
		return 0
	if not _supabase_enabled():
		root = os.path.join(EVENTS_DIR, club_slug)
		total = 0
		for dirpath, _dirs, files in os.walk(root):
			for name in files:
				try:
					total += os.path.getsize(os.path.join(dirpath, name))
				except OSError:
					pass
		return total

	total = 0
	stack = [f"{club_slug}/"]
	visited: set[str] = set()
	while stack:
		prefix = stack.pop()
		if prefix in visited:
			continue
		visited.add(prefix)
		try:
			entries = _supabase_list(prefix)
		except (urlerror.URLError, TimeoutError, OSError, ValueError):
			if strict:
				raise
			continue
		for entry in entries:
			name = entry.get("name")
			if not name:
				continue
			metadata = entry.get("metadata")
			if isinstance(metadata, dict) and metadata.get("size") is not None:
				total += int(metadata["size"])
			else:
				# A folder: recurse into it.
				stack.append(f"{prefix}{name}/")
	return total


def _stored_object_size(object_path: str | None) -> int:
	"""Current size in bytes of one stored object, or 0 if absent/unknown."""
	if not object_path or not _supabase_enabled():
		return 0
	parent, _, leaf = object_path.rpartition("/")
	try:
		for entry in _supabase_list(parent + "/"):
			if entry.get("name") == leaf:
				meta = entry.get("metadata")
				if isinstance(meta, dict) and meta.get("size") is not None:
					return int(meta["size"])
	except (urlerror.URLError, TimeoutError, OSError, ValueError):
		return 0
	return 0


def quota_status(club: dict, incoming_bytes: int,
				 replacing_path: str | None = None) -> tuple[bool, str | None]:
	"""Whether an upload of `incoming_bytes` fits under the club's quota.

	This is a soft cap, not a hard guarantee: two uploads racing can each see room
	and both proceed, so the true usage can overshoot the cap by up to the size of
	the concurrent uploads. Do not present it anywhere as an exact ceiling.

	`replacing_path`, when the upload overwrites an existing object, is that object's
	store path: its current bytes are freed by the upsert, so they are netted out -
	otherwise re-uploading a same-size file near the cap is falsely rejected.
	"""
	cap = int(club.get("quota_bytes", db.DEFAULT_QUOTA_BYTES))
	try:
		used = club_storage_bytes(club["slug"], strict=True)
	except (urlerror.URLError, TimeoutError, OSError, ValueError):
		# Fail closed: if usage cannot be verified, do not admit the upload rather
		# than let a club blow past the shared quota during a Storage outage.
		return False, "Could not verify current storage usage. Please try again shortly."
	net_used = max(0, used - _stored_object_size(replacing_path))
	if net_used + incoming_bytes > cap:
		remaining = max(0, cap - net_used)
		return False, (
			"This upload would exceed the club's storage limit "
			f"({cap // (1024 * 1024)} MB). Used {net_used // (1024 * 1024)} MB, "
			f"{remaining // (1024 * 1024)} MB free. Nothing was uploaded."
		)
	return True, None


def decode_template(data: bytes) -> Image.Image:
	"""
	Decode a template, keeping alpha only when the file actually has it.

	Certificate templates are almost always opaque. Forcing them to RGBA costs 25%
	more memory to hold and 25% more bytes to compress on every single render.
	"""
	image = Image.open(BytesIO(data))
	if image.mode in ("RGBA", "LA") or (image.mode == "P" and "transparency" in image.info):
		return image.convert("RGBA")
	return image.convert("RGB")


def decoded_image_bytes(image: Image.Image) -> int:
	"""Roughly what a decoded image occupies: one byte per band per pixel."""
	return image.width * image.height * len(image.getbands())


def _template_cache_key(slug: str, version, club_slug: str | None = None) -> str:
	"""The single source of truth for the decoded-template cache key. The club is part
	of the key so two clubs sharing an event slug never serve each other's template; a
	legacy (club-less) event keeps its bare key so its cache behaviour is unchanged."""
	return f"{club_slug}/{slug}@{version}" if club_slug else f"{slug}@{version}"


def warm_template_cache_from_bytes(slug: str, version, data: bytes,
								   club_slug: str | None = None) -> None:
	"""Pre-decode a freshly uploaded template into the image cache from the bytes we
	already hold, so the first certificate render skips the storage fetch and decode.
	Best-effort: any failure is swallowed - warming is an optimization, never required
	for correctness, and must never fail an upload."""
	try:
		image = decode_template(data)
		_TEMPLATE_IMAGE_CACHE.put(_template_cache_key(slug, version, club_slug),
								  image, decoded_image_bytes(image))
	except Exception:
		pass


def get_template_image(slug: str, config: dict | None = None,
					   club_slug: str | None = None) -> Image.Image | None:
	"""Decoded template, cached per (club, slug, template version). Returns a copy.

	The club is part of the key: two clubs can own an event with the same slug and
	the same template_version, and a slug-only key would serve one club's template
	for the other's certificate. A legacy (club-less) event keeps its bare key so
	its cache behaviour is unchanged."""
	version = template_version_for(slug, config)
	cache_key = _template_cache_key(slug, version, club_slug)
	cached = _TEMPLATE_IMAGE_CACHE.get(cache_key)
	if cached is None:
		data = load_template_bytes(slug, config, club_slug)
		if data is None:
			return None
		try:
			cached = decode_template(data)
		except Exception:
			return None
		_TEMPLATE_IMAGE_CACHE.put(cache_key, cached, decoded_image_bytes(cached))
	return cached.copy()


def available_font_options() -> list[dict[str, str]]:
	options: list[dict[str, str]] = []
	for key, meta in FONT_OPTIONS.items():
		path = os.path.join(BASE_DIR, "fonts", meta["filename"])
		if not os.path.exists(path):
			continue
		options.append(
			{
				"key": key,
				"label": meta["label"],
				"path": path,
				"css_family": meta["css_family"],
				"css_weight": meta["css_weight"],
			}
		)
	if options:
		return options
	return [
		{
			"key": DEFAULT_FONT_KEY,
			"label": "Default",
			"path": FONT_PATH,
			"css_family": "Arial",
			"css_weight": "700",
		}
	]


def normalize_font_key(value: str | None, fallback: str = DEFAULT_FONT_KEY) -> str:
	candidate = (value or "").strip().lower()
	available = {option["key"] for option in available_font_options()}
	if candidate in available:
		return candidate
	if fallback in available:
		return fallback
	return next(iter(available), DEFAULT_FONT_KEY)


def resolve_font_option(font_key: str | None) -> dict[str, str]:
	normalized_key = normalize_font_key(font_key)
	for option in available_font_options():
		if option["key"] == normalized_key:
			return option
	return available_font_options()[0]


def _normalize_event_style_config(config: dict) -> None:
	config["font_key"] = normalize_font_key(config.get("font_key"), DEFAULT_FONT_KEY)


def _read_event_config_from_file(slug: str) -> dict | None:
	path = _event_config_path(slug)
	if not os.path.exists(path):
		return None
	try:
		with open(path, encoding="utf-8") as f:
			loaded = json.load(f)
		if isinstance(loaded, dict):
			return loaded
	except Exception:
		return None
	return None


def _write_event_config_to_file(slug: str, config: dict) -> None:
	os.makedirs(_event_dir(slug), exist_ok=True)
	with open(_event_config_path(slug), "w", encoding="utf-8") as f:
		json.dump(config, f, indent=2)


def _load_event_config(slug: str) -> dict | None:
	"""Load event config with caching to minimize KV API calls during preview/download."""
	global _EVENT_CONFIG_CACHE

	# Check cache first within TTL
	if slug in _EVENT_CONFIG_CACHE:
		config, cached_at = _EVENT_CONFIG_CACHE[slug]
		if (time.time() - cached_at) < _EVENT_CONFIG_CACHE_TTL_SEC:
			# Copy: callers mutate what they get back, and the cache is shared
			# between requests and between bulk-generation threads.
			return copy.deepcopy(config)

	# Cache miss or expired - reload from source.
	config = None

	# Postgres is the platform's source of truth now; KV is demoted to a fallback
	# for events not yet migrated (Phase 5). A store outage must NOT take down
	# serving of a legacy event that still lives in KV/file, so a DatabaseUnavailable
	# here degrades to the fallback rather than propagating - unlike the auth path,
	# which fails closed. find_event_globally is safe here only while at most one
	# approved club claims the slug, which holds through Phases 1-2.
	try:
		row = repo().find_event_globally(slug)
	except db.DatabaseUnavailable:
		row = None
	if row is not None and isinstance(row.get("config"), dict):
		config = dict(row["config"])
		config.setdefault("slug", row["slug"])
		config.setdefault("name", row["name"])
		config["active"] = row["active"]

	if config is None and _kv_enabled():
		try:
			raw = _kv_get_raw(_event_config_key(slug))
			if raw not in (None, ""):
				loaded = json.loads(raw) if isinstance(raw, str) else raw
				if isinstance(loaded, dict):
					config = loaded
					_ensure_event_slug_registered(slug)
		except (urlerror.URLError, TimeoutError, json.JSONDecodeError, OSError, ValueError):
			pass

	# Fallback to file if KV missed or on error
	if config is None:
		config = _read_event_config_from_file(slug)

	# Cache the result (even if None) to avoid hammering KV on missing configs
	_EVENT_CONFIG_CACHE[slug] = (copy.deepcopy(config), time.time())

	# Limit cache size to prevent memory bloat
	if len(_EVENT_CONFIG_CACHE) > 100:
		_EVENT_CONFIG_CACHE.clear()

	return config


def xlsx_to_csv_text(data: bytes) -> str:
	"""
	Flatten the first worksheet of an .xlsx file to CSV text.

	Everything downstream (validation, dropdowns, bulk generation) already speaks
	CSV, so an upload is converted once here rather than teaching the rest of the
	app about workbooks. The header row sets the column count; blank rows are
	dropped and ragged rows are padded or trimmed to match.
	"""
	from openpyxl import load_workbook

	workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
	try:
		worksheet = workbook.worksheets[0]
		rows: list[list[str]] = []
		for raw_row in worksheet.iter_rows(values_only=True):
			cells = ["" if value is None else str(value).strip() for value in raw_row]
			while cells and cells[-1] == "":
				cells.pop()
			if not cells:
				continue
			rows.append(cells)
	finally:
		workbook.close()

	if not rows:
		return ""

	width = len(rows[0])
	output = StringIO()
	writer = csv.writer(output)
	for row in rows:
		writer.writerow((row + [""] * width)[:width])
	return output.getvalue()


def participant_text_from_upload(data: bytes, ext: str) -> tuple[str | None, str | None]:
	"""Return (csv_text, error). Accepts a .csv or .xlsx upload."""
	if ext == ".xlsx":
		# .xlsx is a zip container; anything else with that extension is a lie.
		if data[:4] != b"PK\x03\x04":
			return None, "That does not look like a valid .xlsx workbook."
		try:
			text = xlsx_to_csv_text(data)
		except Exception:
			return None, "Could not read that workbook. Try re-saving it, or upload a .csv."
		if not text.strip():
			return None, "The first sheet of that workbook is empty."
		return text, None

	# utf-8-sig strips the BOM that Excel writes when it exports CSV.
	return data.decode("utf-8-sig", errors="replace"), None


def _csv_cache_key(slug: str, club_slug: str | None = None) -> str:
	"""Participant caches are keyed by (club, slug). A legacy csi-aseb event keeps
	the bare slug so its cache behaviour is unchanged; two clubs sharing an event
	slug get distinct keys, so one club can never be served the other's roster."""
	return f"{club_slug}/{slug}" if club_slug else slug


def _read_event_csv_from_file(slug: str, club_slug: str | None = None) -> str | None:
	path = os.path.join(_scoped_event_dir(slug, club_slug), "data.csv") if club_slug else _event_csv_path(slug)
	if not os.path.exists(path):
		return None
	try:
		with open(path, newline="", encoding="utf-8") as f:
			return f.read()
	except OSError:
		return None


def load_event_csv_text(slug: str, club_slug: str | None = None) -> str | None:
	"""
	Participant CSV text, cached briefly because one page view reads it repeatedly.

	Supabase is the store; KV is read only (legacy events) and the local file is the
	no-storage fallback. A club event reads from its own <club>/<event> prefix and
	caches under a club-scoped key, so one club never sees another's participants.
	"""
	key = _csv_cache_key(slug, club_slug)
	cached = _EVENT_CSV_CACHE.get(key)
	if cached is not None and (time.time() - cached[1]) < _EVENT_CSV_CACHE_TTL_SEC:
		return cached[0]

	content: str | None = None
	if _supabase_enabled():
		try:
			raw_bytes = _supabase_download(_participants_object_path(slug, club_slug=club_slug))
			if raw_bytes:
				content = raw_bytes.decode("utf-8-sig", errors="replace")
		except (urlerror.URLError, TimeoutError, OSError, ValueError):
			pass
	# KV holds only legacy (club-less) events; a club event never consults it.
	if content is None and club_slug is None and _kv_enabled():
		try:
			raw = _kv_get_raw(_event_csv_key(slug))
			if isinstance(raw, str):
				content = raw
		except (urlerror.URLError, TimeoutError, OSError, ValueError):
			pass
	if content is None:
		content = _read_event_csv_from_file(slug, club_slug)

	_EVENT_CSV_CACHE[key] = (content, time.time())
	if len(_EVENT_CSV_CACHE) > 100:
		_EVENT_CSV_CACHE.clear()
	return content


def _write_event_csv_to_file(slug: str, content: str) -> None:
	os.makedirs(_event_dir(slug), exist_ok=True)
	with open(_event_csv_path(slug), "w", encoding="utf-8", newline="") as f:
		f.write(content)


def _load_kv_event_index() -> list[str]:
	if not _kv_enabled():
		return []
	try:
		raw = _kv_get_raw(KV_EVENT_INDEX_KEY)
		if raw in (None, ""):
			return []
		loaded = json.loads(raw) if isinstance(raw, str) else raw
		if not isinstance(loaded, list):
			return []
		seen: set[str] = set()
		result: list[str] = []
		for value in loaded:
			slug = str(value)
			if safe_slug(slug) and slug not in seen:
				seen.add(slug)
				result.append(slug)
		return result
	except (urlerror.URLError, TimeoutError, json.JSONDecodeError, OSError, ValueError):
		return []


def _save_kv_event_index(slugs: list[str]) -> None:
	if not _kv_enabled():
		return
	encoded = json.dumps(slugs, separators=(",", ":"))
	_kv_set_raw(KV_EVENT_INDEX_KEY, encoded)


def _register_event_slug(slug: str) -> None:
	if not _kv_enabled():
		return
	try:
		slugs = _load_kv_event_index()
		if slug not in slugs:
			slugs.append(slug)
			_save_kv_event_index(slugs)
	except (urlerror.URLError, TimeoutError, OSError, ValueError):
		return


def _ensure_event_slug_registered(slug: str) -> None:
	"""Best-effort index repair for cases where config exists but index is stale."""
	if not _kv_enabled() or not safe_slug(slug):
		return
	_register_event_slug(slug)


def _unregister_event_slug(slug: str) -> None:
	if not _kv_enabled():
		return
	try:
		slugs = [value for value in _load_kv_event_index() if value != slug]
		_save_kv_event_index(slugs)
	except (urlerror.URLError, TimeoutError, OSError, ValueError):
		return


def _event_exists(slug: str) -> bool:
	if not safe_slug(slug):
		return False
	# Slugs explicitly marked deleted are considered available for recreation.
	if _event_state(slug).get("deleted", False):
		return False
	return _load_event_config(slug) is not None


def _event_csv_exists(slug: str) -> bool:
	return load_event_csv_text(slug) is not None


def _all_event_slugs() -> list[str]:
	seen: set[str] = set()
	result: list[str] = []
	if os.path.isdir(EVENTS_DIR):
		for slug in os.listdir(EVENTS_DIR):
			if safe_slug(slug) and slug not in seen:
				seen.add(slug)
				result.append(slug)
	# Include event state keys so state/index drift does not hide valid events.
	for slug in _load_event_states().keys():
		if safe_slug(slug) and slug not in seen:
			seen.add(slug)
			result.append(slug)
	for slug in _load_kv_event_index():
		if slug not in seen:
			seen.add(slug)
			result.append(slug)
	return result


def _apply_profile_overrides(config: dict) -> None:
	mode = config.get("mode") or config.get("profile")
	if not mode:
		return

	profiles_path = os.path.join(BASE_DIR, "profiles.yaml")
	if not os.path.exists(profiles_path):
		return

	try:
		with open(profiles_path, "r", encoding="utf-8") as f:
			profiles_data = yaml.safe_load(f)
			if not isinstance(profiles_data, dict):
				return

			profiles = profiles_data.get("profiles", {})
			profile_config = profiles.get(mode)

			if isinstance(profile_config, dict):
				# Inherit keys if missing from the event config
				for key in ["text_x", "text_y", "font_size", "font_color", "font_key"]:
					if key in profile_config and key not in config:
						config[key] = profile_config[key]
	except Exception as e:
		print(f"Error loading profile {mode}: {e}")

def load_event(slug: str, states: dict[str, dict] | None = None) -> dict | None:
	if not safe_slug(slug):
		return None
	state = _event_state(slug, states)
	if state.get("deleted", False):
		return None
	config = _load_event_config(slug)
	if config is None:
		return None
	_apply_profile_overrides(config)
	_normalize_event_style_config(config)
	if "active" in state:
		config["active"] = bool(state.get("active"))
	return config


def save_event_config(slug: str, config: dict) -> None:
	_normalize_event_style_config(config)
	_write_event_config_to_file(slug, config)
	# Invalidate cache so next load gets the new config
	global _EVENT_CONFIG_CACHE
	_EVENT_CONFIG_CACHE.pop(slug, None)
	if _kv_enabled():
		try:
			_kv_set_raw(_event_config_key(slug), json.dumps(config, separators=(",", ":")))
			_register_event_slug(slug)
		except (urlerror.URLError, TimeoutError, OSError, ValueError):
			return


def save_event_csv(slug: str, content: str, source: tuple[bytes, str] | None = None,
				   club_slug: str | None = None) -> None:
	"""
	Persist the participant list.

	`source` is the original upload as (bytes, extension); when it is a workbook
	it is kept alongside the derived CSV so the organiser can download exactly
	what they uploaded.
	"""
	local_dir = _scoped_event_dir(slug, club_slug)
	try:
		os.makedirs(local_dir, exist_ok=True)
		with open(os.path.join(local_dir, "data.csv"), "w", encoding="utf-8", newline="") as f:
			f.write(content)
	except OSError:
		if not (_supabase_enabled() or (_kv_enabled() and not club_slug)):
			raise
	_EVENT_CSV_CACHE.pop(_csv_cache_key(slug, club_slug), None)
	_PARTICIPANT_DATASET_CACHE.pop(_csv_cache_key(slug, club_slug), None)

	if _supabase_enabled():
		_supabase_upload(
			_participants_object_path(slug, club_slug=club_slug),
			content.encode("utf-8"),
			PARTICIPANT_CONTENT_TYPES[".csv"],
		)
		if source is not None and source[1] != ".csv":
			raw_bytes, ext = source
			try:
				_supabase_upload(
					_participants_object_path(slug, f"source{ext}", club_slug=club_slug),
					raw_bytes,
					PARTICIPANT_CONTENT_TYPES.get(ext, "application/octet-stream"),
				)
			except (urlerror.URLError, TimeoutError, OSError, ValueError):
				# Keeping the original is a convenience, not a requirement.
				pass
		if not club_slug:
			_register_event_slug(slug)
		return

	# No Supabase and a legacy event: fall back to KV so old deployments still
	# work. Club events never touch KV - it is demoted and single-club.
	if _kv_enabled() and not club_slug:
		try:
			_kv_set_raw(_event_csv_key(slug), content)
			_register_event_slug(slug)
		except (urlerror.URLError, TimeoutError, OSError, ValueError):
			return


def delete_event_storage(slug: str, club_slug: str | None = None) -> None:
	local_dir = _scoped_event_dir(slug, club_slug)
	if os.path.isdir(local_dir):
		shutil.rmtree(local_dir)
	delete_event_objects(slug, club_slug)
	# Invalidate caches for the deleted event
	global _EVENT_CONFIG_CACHE
	_EVENT_CONFIG_CACHE.pop(slug, None)
	_EVENT_CSV_CACHE.pop(_csv_cache_key(slug, club_slug), None)
	_PARTICIPANT_DATASET_CACHE.pop(_csv_cache_key(slug, club_slug), None)
	if _kv_enabled():
		try:
			_kv_delete_key(_event_config_key(slug))
			_kv_delete_key(_event_csv_key(slug))
			_unregister_event_slug(slug)
		except (urlerror.URLError, TimeoutError, OSError, ValueError):
			return


def load_all_events(active_only: bool = False) -> list[dict]:
	events = []
	states = _load_event_states()
	for slug in _all_event_slugs():
		config = load_event(slug, states)
		if config is None:
			continue
		if active_only and not config.get("active", False):
			continue
		events.append(config)
	events.sort(key=lambda e: e.get("name", "").lower())
	return events


def normalize_value(value: str) -> str:
	return (value or "").strip().lower()


def parse_custom_fields(form_values: list[str]) -> list[str]:
	fields: list[str] = []
	seen: set[str] = set()
	for value in form_values:
		for token in (value or "").split(","):
			field = normalize_value(token)
			if field and field not in seen:
				seen.add(field)
				fields.append(field)
	return fields


def csv_headers(slug: str, club_slug: str | None = None) -> list[str]:
	return list(load_participant_dataset(slug, club_slug).columns)


# ─── Participant dataset ──────────────────────────────────────────────────────
#
# Every page view used to re-read and re-parse the CSV several times over: the
# team list, each dropdown column, then again on validation. At 1000 arrivals in
# minutes that is the CSV parsed thousands of times. This parses it once per
# distinct CSV content and hands every loader the same structure.
#
# It also draws the line the multi-field feature needs: values are kept BOTH
# verbatim (what prints on the certificate - "Nimbus", "07") and normalized
# (lowercased, for case-insensitive matching). Resolving a csv field off the
# normalized copy would print "nimbus" and turn "07" matching into a display bug,
# which is exactly what the plan forbids.

# Keyed by slug, validated by a hash of the CSV text, so replacing the CSV yields
# a new hash and a fresh parse with no manual invalidation needed. Upload paths
# still evict eagerly (below) so the change shows before the text cache's TTL.
_PARTICIPANT_DATASET_CACHE: dict[str, tuple[str, "ParticipantDataset"]] = {}


class ParticipantDataset:
	"""One parse of an event's participant CSV, verbatim values preserved.

	raw_rows[i] and norm_rows[i] are the same row: raw for display, normalized for
	matching. Header keys are normalized in both, so a column is looked up the same
	way regardless of how the CSV cased its header.
	"""

	__slots__ = ("columns", "raw_rows", "norm_rows", "_column_values")

	def __init__(self, columns: list[str], raw_rows: list[dict[str, str]],
				 norm_rows: list[dict[str, str]]):
		self.columns = columns
		self.raw_rows = raw_rows
		self.norm_rows = norm_rows
		self._column_values: dict[str, list[str]] = {}

	def column_values(self, column: str) -> list[str]:
		"""Distinct verbatim values of a column, first-seen order. Memoized."""
		col = normalize_value(column)
		if col in self._column_values:
			return self._column_values[col]
		seen: set[str] = set()
		values: list[str] = []
		for raw, norm in zip(self.raw_rows, self.norm_rows):
			key = norm.get(col, "")
			raw_value = raw.get(col, "")
			if raw_value and key not in seen:
				seen.add(key)
				values.append(raw_value)
		self._column_values[col] = values
		return values

	def rows_matching(self, key: dict[str, str]) -> list[dict[str, str]]:
		"""Raw rows whose normalized values equal every (column, value) in `key`.

		`key` values are compared normalized, so matching stays case-insensitive
		while the rows returned keep their verbatim values for display.
		"""
		norm_key = {normalize_value(c): normalize_value(v) for c, v in key.items()}
		out: list[dict[str, str]] = []
		for raw, norm in zip(self.raw_rows, self.norm_rows):
			if all(norm.get(c, "") == v for c, v in norm_key.items()):
				out.append(raw)
		return out


_EMPTY_DATASET = ParticipantDataset([], [], [])


def _parse_participant_dataset(content: str) -> ParticipantDataset:
	reader = csv.DictReader(content.splitlines())
	columns = [normalize_value(h) for h in (reader.fieldnames or []) if normalize_value(h)]
	raw_rows: list[dict[str, str]] = []
	norm_rows: list[dict[str, str]] = []
	for row in reader:
		raw: dict[str, str] = {}
		norm: dict[str, str] = {}
		for header, value in row.items():
			col = normalize_value(header)
			if not col:
				continue
			# Verbatim except for a surrounding-whitespace trim, so a stray space
			# in a cell does not defeat matching or print as leading padding.
			raw[col] = (value or "").strip()
			norm[col] = normalize_value(value or "")
		raw_rows.append(raw)
		norm_rows.append(norm)
	return ParticipantDataset(columns, raw_rows, norm_rows)


def load_participant_dataset(slug: str, club_slug: str | None = None) -> ParticipantDataset:
	"""Parsed participant CSV for an event, cached per distinct CSV content and per
	club, so two clubs sharing a slug never share a dataset."""
	content = load_event_csv_text(slug, club_slug)
	if content is None:
		return _EMPTY_DATASET
	key = _csv_cache_key(slug, club_slug)
	fingerprint = hashlib.sha256(content.encode("utf-8", "replace")).hexdigest()[:16]
	cached = _PARTICIPANT_DATASET_CACHE.get(key)
	if cached is not None and cached[0] == fingerprint:
		return cached[1]
	dataset = _parse_participant_dataset(content)
	_PARTICIPANT_DATASET_CACHE[key] = (fingerprint, dataset)
	if len(_PARTICIPANT_DATASET_CACHE) > 100:
		_PARTICIPANT_DATASET_CACHE.clear()
		_PARTICIPANT_DATASET_CACHE[key] = (fingerprint, dataset)
	return dataset


def load_csv_rows(slug: str, club_slug: str | None = None) -> list[dict[str, str]]:
	"""Normalized rows, for matching. Same shape as before; now a single parse."""
	return load_participant_dataset(slug, club_slug).norm_rows


def required_headers_for_validation(validation_type: str, custom_fields: list[str]) -> set[str]:
	if validation_type == "player_team":
		return {"player", "team"}
	if validation_type == "name_only":
		return {"name"}
	if validation_type == "email":
		return {"email"}
	if validation_type == "custom":
		return set(custom_fields)
	return set()


def build_custom_form_fields(slug: str, custom_fields: list[str], custom_dropdown_fields: list[str] | None = None, club_slug: str | None = None) -> list[dict]:
	result: list[dict] = []
	dropdown_set = set(custom_dropdown_fields or [])
	for field in custom_fields:
		key = re.sub(r"[^a-z0-9]+", "_", field).strip("_")
		if not key:
			continue
		is_dropdown = field in dropdown_set
		result.append(
			{
				"column": field,
				"key": key,
				"label": field.replace("_", " ").title(),
				"is_dropdown": is_dropdown,
				"options": load_unique_column_values(slug, field, club_slug) if is_dropdown else [],
			}
		)
	return result


def validation_prompt_for_type(validation_type: str) -> str:
	if validation_type == "email":
		return "Registration email"
	if validation_type == "badge_id":
		return "Roll number"
	return "Registration name"


def participant_input_fields(config: dict) -> tuple:
	"""(primary_id, extra_fields) for the participant form.

	The primary input field is the certificate's `name` field if it has one, else
	its first input field; the big `cert_name` box on the event page fills it. Any
	other input field gets its own labelled box, posted as field_<id>. csv/static
	fields never appear here - they resolve server-side.
	"""
	input_fields = [f for f in normalize_fields(config) if f["source"] == "input"]
	if not input_fields:
		return None, []
	primary_id = next((f["id"] for f in input_fields if f["id"] == "name"), input_fields[0]["id"])
	extra = [{"id": f["id"], "label": f["label"]} for f in input_fields if f["id"] != primary_id]
	return primary_id, extra


def event_form_context(config: dict, slug: str, error: str | None = None,
					   download_action: str | None = None, club_slug: str | None = None) -> dict:
	validation_type = config.get("validation_type", "player_team")
	custom_fields = config.get("custom_fields", [])
	custom_dropdown_fields = config.get("custom_dropdown_fields", [])
	validation_prompt = validation_prompt_for_type(validation_type)
	registration_placeholder = "BL.SC.U4AIExxxxx" if validation_type == "badge_id" else f"Enter {validation_prompt.lower()}"
	primary_id, extra_input_fields = participant_input_fields(config)
	name_label = "Name to print on the certificate"
	for field in normalize_fields(config):
		if field["id"] == primary_id:
			name_label = field["label"] or name_label
			break
	return {
		"event": config,
		"teams": load_team_names(slug, club_slug) if validation_type == "player_team" else [],
		"custom_form_fields": build_custom_form_fields(slug, custom_fields, custom_dropdown_fields, club_slug),
		"validation_prompt": validation_prompt,
		"registration_placeholder": registration_placeholder,
		"name_label": name_label,
		"extra_input_fields": extra_input_fields,
		# The "name to print" box only makes sense when a field is filled by participant
		# input. When the name is read from the participants file (csv) it is pulled from
		# the matched row, so the box is neither shown nor required.
		"has_name_input": primary_id is not None,
		"download_action": download_action or url_for("download_certificate", slug=slug),
		"error": error,
	}


def load_valid_participants(slug: str, club_slug: str | None = None) -> set[tuple[str, str]]:
	participants: set[tuple[str, str]] = set()
	for row in load_participant_dataset(slug, club_slug).norm_rows:
		player, team = row.get("player", ""), row.get("team", "")
		if player and team:
			participants.add((player, team))
	return participants


def load_valid_names(slug: str, club_slug: str | None = None) -> set[str]:
	return {row.get("name", "") for row in load_participant_dataset(slug, club_slug).norm_rows
			if row.get("name", "")}


def load_team_names(slug: str, club_slug: str | None = None) -> list[str]:
	return sorted(load_participant_dataset(slug, club_slug).column_values("team"),
				  key=lambda v: v.lower())


def load_unique_column_values(slug: str, column: str, club_slug: str | None = None) -> list[str]:
	return load_participant_dataset(slug, club_slug).column_values(column)


def validate_participant_submission(slug: str, config: dict, form_data, club_slug: str | None = None) -> tuple[list[dict[str, str]] | None, str | None]:
	"""Validate a submission and return the rows it matched.

	Returns ``(matched_rows, None)`` on success and ``(None, error)`` on failure.
	``matched_rows`` holds every raw (verbatim) CSV row the submission matched, so a
	``csv`` field can be resolved against them and, when a coarse key like team-only
	matches several people, disagreement can be detected instead of silently taking
	the first row. ``validation_type == "none"`` performs no match and returns
	``([], None)`` - a valid submission with no row behind it.

	The single ``None``-vs-list return is what separates failure from a legitimately
	empty match: ``None`` is always an error, a list (even empty) is always success.
	"""
	validation_type = config.get("validation_type", "player_team")
	custom_fields: list[str] = config.get("custom_fields", [])

	if validation_type == "none":
		return [], None

	dataset = load_participant_dataset(slug, club_slug)

	if validation_type == "player_team":
		registration_name = normalize_value(form_data.get("registration_name", ""))
		team_name = normalize_value(form_data.get("team_name", ""))
		if not registration_name or not team_name:
			return None, "Please fill all fields."
		matched = dataset.rows_matching({"player": registration_name, "team": team_name})
		if not matched:
			return None, "Invalid player or team name."
		return matched, None

	if validation_type == "name_only":
		registration_name = normalize_value(form_data.get("registration_name", ""))
		if not registration_name:
			return None, "Please fill all fields."
		matched = dataset.rows_matching({"name": registration_name})
		if not matched:
			return None, "Name not found in participant list."
		return matched, None

	if validation_type == "email":
		registration_email = normalize_value(form_data.get("registration_name", ""))
		if not registration_email:
			return None, "Please fill all fields."
		matched = dataset.rows_matching({"email": registration_email})
		if not matched:
			return None, "Email not found in participant list."
		return matched, None

	if validation_type == "badge_id":
		registration_id = normalize_value(form_data.get("registration_name", ""))
		if not registration_id:
			return None, "Please fill all fields."
		# An id may live under any of these headers, so this is an OR across columns
		# rather than a single-key match - kept exactly as before.
		id_columns = ("roll_no", "id", "badge_id", "badge_number")
		matched = [raw for raw, norm in zip(dataset.raw_rows, dataset.norm_rows)
				   if any(norm.get(col, "") == registration_id for col in id_columns)]
		if not matched:
			return None, "Roll No not found in participant list."
		return matched, None

	if validation_type == "custom":
		if not custom_fields:
			return None, "Custom validation fields are not configured by admin."
		form_fields = build_custom_form_fields(slug, custom_fields, club_slug=club_slug)
		expected: dict[str, str] = {}
		for field in form_fields:
			value = normalize_value(form_data.get(f"custom_{field['key']}", ""))
			if not value:
				return None, "Please fill all fields."
			expected[field["column"]] = value
		matched = dataset.rows_matching(expected)
		if not matched:
			return None, "Details not found in participant list."
		return matched, None

	return None, "Unsupported validation type configured for this event."


# ─── Certificate helpers ──────────────────────────────────────────────────────

def get_font(size: int, font_key: str | None = None) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
	"""Load font with caching to avoid repeated disk I/O on every render."""
	cache_key = f"{size}:{font_key or DEFAULT_FONT_KEY}"
	cached = _cache_get(_FONT_CACHE, cache_key)
	if cached is not None:
		return cached

	font_option = resolve_font_option(font_key)
	try:
		if os.path.exists(font_option["path"]):
			font = ImageFont.truetype(font_option["path"], size=size)
		else:
			font = ImageFont.truetype("arial.ttf", size=size)
	except Exception:
		font = ImageFont.load_default()

	_cache_put(_FONT_CACHE, cache_key, font, _FONT_CACHE_MAX)
	return font


def _cert_metadata_path(cert_id: str) -> str:
	"""Location of a pre-token certificate record. Retained for legacy links only."""
	return os.path.join(GENERATED_DIR, f"{cert_id}.json")


def certificate_render_settings(config: dict) -> dict:
	"""The subset of an event config that actually affects the rendered image."""
	return {
		"text_x": config.get("text_x", 1789),
		"text_y": config.get("text_y", 1440),
		"font_size": config.get("font_size", 100),
		"font_color": config.get("font_color", [50, 34, 24]),
		"font_key": normalize_font_key(config.get("font_key"), DEFAULT_FONT_KEY),
	}


# --- Certificate fields -------------------------------------------------------
#
# A certificate carries up to five text fields, each with its own placement and
# type. A field's `source` decides where its printed value comes from and, more
# importantly, who controls it:
#
#   input  - the participant types it (as the name always has been)
#   csv    - read from the row they validated against; they cannot change it
#   static - a fixed string the club set on the event
#
# Only `input` is participant-controlled. A `position` or `points` field must be
# `csv` or `static`, or a participant awards themselves whatever they like - that
# distinction is the whole point of the feature, enforced at resolve time below.

MAX_FIELDS = 5
FIELD_SOURCES = ("input", "csv", "static")
FIELD_ALIGNS = ("left", "center", "right")
FIELD_OVERFLOWS = ("shrink", "truncate")
_ALIGN_ANCHORS = {"left": "lm", "center": "mm", "right": "rm"}
_TRUNCATE_ELLIPSIS = "\u2026"


def _coerce_color(raw) -> list[int]:
	try:
		channels = [max(0, min(255, int(c))) for c in raw][:3]
	except (TypeError, ValueError):
		return [50, 34, 24]
	return channels if len(channels) == 3 else [50, 34, 24]


def _parse_int_value(value, fallback: int) -> int:
	try:
		return max(0, int(value))
	except (TypeError, ValueError):
		return fallback


def _legacy_name_field(config: dict) -> dict:
	"""Synthesize the single input field an event carried before fields existed.

	`max_width` stays None so no measurement happens and the draw is the exact
	anchored call the old renderer made - existing events must stay byte-identical
	until a club edits them.
	"""
	return {
		"id": "name",
		"label": "Name",
		"source": "input",
		"column": "",
		"value": "",
		"x": _parse_int_value(config.get("text_x"), 1789),
		"y": _parse_int_value(config.get("text_y"), 1440),
		"font_size": _parse_int_value(config.get("font_size"), 100),
		"font_color": _coerce_color(config.get("font_color", [50, 34, 24])),
		"font_key": normalize_font_key(config.get("font_key"), DEFAULT_FONT_KEY),
		"align": "center",
		"max_width": None,
		"overflow": "shrink",
		"depends_on": None,
	}


def _normalize_one_field(raw: dict, index: int, used_ids: set) -> dict | None:
	if not isinstance(raw, dict):
		return None
	source = str(raw.get("source", "")).strip().lower()
	if source not in FIELD_SOURCES:
		source = "input"

	field_id = re.sub(r"[^a-z0-9_]+", "_", str(raw.get("id", "")).strip().lower()).strip("_")
	if not field_id:
		field_id = "field_%d" % (index + 1)
	while field_id in used_ids:
		field_id = "%s_%d" % (field_id, index + 1)
	used_ids.add(field_id)

	align = str(raw.get("align", "center")).strip().lower()
	if align not in FIELD_ALIGNS:
		align = "center"
	overflow = str(raw.get("overflow", "shrink")).strip().lower()
	if overflow not in FIELD_OVERFLOWS:
		overflow = "shrink"

	max_width_raw = raw.get("max_width")
	max_width = None
	if max_width_raw not in (None, "", 0, "0"):
		parsed = _parse_int_value(max_width_raw, 0)
		max_width = parsed if parsed > 0 else None

	# depends_on is reserved for the dependent-dropdown feature (0.5b). It is
	# accepted and carried through now so the schema is genuinely frozen, and
	# ignored everywhere in 0.5a. Kept only when it names another field.
	depends_on_raw = raw.get("depends_on")
	depends_on = str(depends_on_raw).strip().lower() or None if depends_on_raw else None

	return {
		"id": field_id,
		"label": str(raw.get("label", "") or field_id).strip() or field_id,
		"source": source,
		"column": normalize_value(str(raw.get("column", ""))) if source == "csv" else "",
		"value": str(raw.get("value", "")) if source == "static" else "",
		"x": _parse_int_value(raw.get("x"), 0),
		"y": _parse_int_value(raw.get("y"), 0),
		"font_size": max(1, _parse_int_value(raw.get("font_size"), 100)),
		"font_color": _coerce_color(raw.get("font_color", [50, 34, 24])),
		"font_key": normalize_font_key(raw.get("font_key"), DEFAULT_FONT_KEY),
		"align": align,
		"max_width": max_width,
		"overflow": overflow,
		"depends_on": depends_on,
	}


def normalize_fields(config: dict) -> list:
	"""The event's fields as a clean, capped list, always render-safe.

	An event with no `fields` synthesizes one input field from the legacy scalars,
	so old configs keep rendering byte-identically. A stored list is normalized and
	capped at MAX_FIELDS - a hand-edited or migrated config with more is truncated
	rather than trusted; the editor (0.5b) rejects an over-long list at save with a
	message. `depends_on` is validated as reserved and otherwise left inert.
	"""
	raw_fields = config.get("fields")
	if not isinstance(raw_fields, list) or not raw_fields:
		return [_legacy_name_field(config)]

	fields = []
	used_ids = set()
	for index, raw in enumerate(raw_fields):
		if len(fields) >= MAX_FIELDS:
			break
		normalized = _normalize_one_field(raw, index, used_ids)
		if normalized is not None:
			fields.append(normalized)
	# Drop any depends_on pointing at a field that did not survive normalization.
	valid_ids = {f["id"] for f in fields}
	for field in fields:
		if field["depends_on"] not in valid_ids:
			field["depends_on"] = None
	return fields or [_legacy_name_field(config)]


def _json_for_script(value) -> str:
	"""json.dumps escaped so it is safe to embed inside a <script> block: a CSV cell
	containing "</script>" or a lone "<" cannot break out of the tag."""
	return (json.dumps(value)
			.replace("<", "\u003c").replace(">", "\u003e").replace("&", "\u0026"))


def validate_fields_payload(slug: str, raw_fields, club_slug: str | None = None) -> tuple:
	"""Validate a fields list submitted by the editor. (fields, None) or (None, error).

	This is the server-side gate the plan calls for: it REJECTS rather than silently
	coerces, so a broken or over-long list never reaches storage. It is the third of
	the three cap-enforcement points (editor disables Add at five, mint refuses to
	resolve more than five, and this rejects a longer list on save).
	"""
	if not isinstance(raw_fields, list):
		return None, "No fields were submitted."
	if not raw_fields:
		return None, "A certificate needs at least one field."
	if len(raw_fields) > MAX_FIELDS:
		return None, "A certificate can have at most %d fields." % MAX_FIELDS

	columns = set(csv_headers(slug, club_slug))
	fields = []
	used_ids = set()
	for index, raw in enumerate(raw_fields):
		if not isinstance(raw, dict):
			return None, "Field %d is not valid." % (index + 1)
		source = str(raw.get("source", "")).strip().lower()
		if source not in FIELD_SOURCES:
			return None, "Field %d has an unknown source." % (index + 1)
		normalized = _normalize_one_field(raw, index, used_ids)
		label = normalized["label"]
		if source == "csv":
			if not normalized["column"]:
				return None, 'The field "%s" needs a column from the participants file.' % label
			if columns and normalized["column"] not in columns:
				return None, ('The field "%s" refers to a column ("%s") that is not in '
							  "the uploaded participants file." % (label, normalized["column"]))
			if not columns:
				return None, ('The field "%s" reads from the participants file, but no '
							  "file has been uploaded yet." % label)
		fields.append(normalized)
	# depends_on can only point at a field that exists in this same list.
	valid_ids = {f["id"] for f in fields}
	for field in fields:
		if field["depends_on"] not in valid_ids:
			field["depends_on"] = None
	return fields, None


def resolve_field_values(fields: list, matched_rows: list,
						 inputs: dict) -> tuple:
	"""Resolve every field to its printed string. (values, None) or (None, error).

	`csv` values come verbatim from the matched rows; when a coarse key matched
	several people and a `csv` field disagrees between them, the field is ambiguous
	and this refuses rather than silently printing the first row's value. `input`
	values come only from `inputs` - a csv or static field can never be fed from
	participant input, which is the security boundary the feature exists to draw.
	"""
	values = {}
	for field in fields[:MAX_FIELDS]:
		source = field["source"]
		if source == "input":
			values[field["id"]] = (inputs.get(field["id"], "") or "").strip()
		elif source == "static":
			values[field["id"]] = field["value"]
		else:  # csv
			column = field["column"]
			distinct = {row.get(column, "") for row in matched_rows}
			if len(distinct) > 1:
				return None, (
					"Please also select your name so \u201c%s\u201d can be filled in "
					"\u2014 it differs between the matching entries." % field["label"]
				)
			values[field["id"]] = matched_rows[0].get(column, "") if matched_rows else ""
	return values, None


def resolved_fields_for_render(config: dict, values: dict) -> list:
	"""Pair each normalized field with the value the token carried for it."""
	fields = normalize_fields(config)
	for field in fields:
		field["text"] = values.get(field["id"], "")
	return fields


def display_name(values: dict) -> str:
	"""The name to show in headings and download filenames: the name field, or the
	first non-empty value if this event has no field called `name`."""
	if values.get("name"):
		return values["name"]
	return next((v for v in values.values() if v), "")


def _fit_text(draw, text: str, field: dict, font):
	"""Apply a field's max_width: shrink the font until it fits, or truncate.

	Returns the text and font to actually draw. No-op when the text already fits or
	the field has no max_width, so an unconstrained field costs no measurement.
	"""
	max_width = field["max_width"]
	if not max_width or not text:
		return text, font
	if draw.textlength(text, font=font) <= max_width:
		return text, font
	if field["overflow"] == "truncate":
		clipped = text
		while len(clipped) > 1 and draw.textlength(clipped + _TRUNCATE_ELLIPSIS, font=font) > max_width:
			clipped = clipped[:-1]
		return clipped + _TRUNCATE_ELLIPSIS, font
	# shrink
	size = field["font_size"]
	while size > 8 and draw.textlength(text, font=font) > max_width:
		size = max(8, int(size * 0.94))
		font = get_font(size, field["font_key"])
	return text, font


def draw_fields_on_image(image: Image.Image, fields: list) -> None:
	"""Draw every field (each carrying its resolved `text`) onto the template.

	A single input field with center align and no max_width takes the same anchored
	draw call the old single-name renderer used, so a legacy event is byte-identical.
	"""
	draw = ImageDraw.Draw(image)
	for field in fields:
		text = field.get("text", "")
		if not text:
			continue
		font = get_font(field["font_size"], field["font_key"])
		color = tuple(field["font_color"])
		position = (field["x"], field["y"])
		if isinstance(font, ImageFont.FreeTypeFont):
			text, font = _fit_text(draw, text, field, font)
			draw.text(position, text, fill=color, font=font,
					  anchor=_ALIGN_ANCHORS.get(field["align"], "mm"))
		else:
			# load_default() is the no-TTF fallback and does not support anchor=.
			draw.text(position, text, fill=color, font=font)


def normalize_download_format(value: str | None,
							  fallback: str = DEFAULT_DOWNLOAD_FORMAT) -> str:
	"""Coerce a stored or submitted download format to one we actually encode."""
	candidate = (value or "").strip().lower()
	if candidate in ("jpeg", "jpg"):
		return "jpeg"
	if candidate == "png":
		return "png"
	return fallback if fallback in DOWNLOAD_FORMATS else DEFAULT_DOWNLOAD_FORMAT


def encode_certificate(image: Image.Image, variant: str = "download",
					   download_format: str | None = None) -> tuple[bytes, str]:
	"""
	Encode a rendered certificate. Returns (bytes, mimetype).

	"preview" is what the browser shows in an <img>: downscaled and JPEG, because
	nobody looks at 3508 px on a phone and the full-size PNG is ~2.5 MB.
	"download" is the real artifact, at full resolution, in the event's format.

	A template with real alpha is always PNG whatever the event asks for: JPEG has
	no alpha channel, so the transparency would come back as a black background.
	"""
	has_alpha = image.mode in ("RGBA", "LA")
	if variant == "preview":
		if image.width > PREVIEW_MAX_WIDTH:
			# BOX is the cheapest filter that still looks clean at a ~3x reduction.
			image.thumbnail((PREVIEW_MAX_WIDTH, PREVIEW_MAX_WIDTH), Image.BOX)
		if not has_alpha:
			output = BytesIO()
			image.save(output, format="JPEG", quality=PREVIEW_JPEG_QUALITY)
			return output.getvalue(), "image/jpeg"
	elif not has_alpha and normalize_download_format(download_format) == "jpeg":
		output = BytesIO()
		image.save(output, format="JPEG", quality=DOWNLOAD_JPEG_QUALITY,
				   subsampling=DOWNLOAD_JPEG_SUBSAMPLING)
		return output.getvalue(), "image/jpeg"
	output = BytesIO()
	image.save(output, format="PNG", compress_level=DOWNLOAD_PNG_COMPRESS_LEVEL)
	return output.getvalue(), "image/png"


def render_certificate(slug: str, cert_values, config: dict,
					   variant: str = "download",
					   tenant_key: str | None = None,
					   club_slug: str | None = None) -> tuple[bytes, str, str] | None:
	"""
	Render a certificate on demand and return (image_bytes, etag, mimetype).

	`cert_values` is the field values to print: a dict {field_id: text}, or a bare
	string for the common single-name case (treated as {"name": text}). Nothing is
	written to disk - template plus config plus these values reproduce the image on
	any worker.

	The etag covers the template version, the download format, the variant, and
	EVERY field's placement, type, and resolved value - so moving or restyling one
	field, or changing another field's value, all produce a new etag instead of a
	stale hit that would serve the other fields wrong.

	Raises RenderCapacityError if no render slot comes free in time. Returns None
	only when the event has no usable template - the caller must not conflate the
	two: one is a 503, the other a 404.
	"""
	values = {"name": cert_values} if isinstance(cert_values, str) else dict(cert_values or {})
	fields = resolved_fields_for_render(config, values)
	# The requested format, not the resolved one - resolving needs the decoded
	# image, and the request is what varies. Alpha is already covered by the
	# template version, so the pair still determines the bytes uniquely.
	download_format = normalize_download_format(config.get("download_format"))
	fingerprint_data = {
		"slug": slug,
		"template": template_version_for(slug, config),
		"variant": variant,
		"format": download_format if variant == "download" else "preview",
		"fields": [
			{
				"id": f["id"], "text": f["text"], "x": f["x"], "y": f["y"],
				"font_size": f["font_size"], "font_color": f["font_color"],
				"font_key": f["font_key"], "align": f["align"],
				"max_width": f["max_width"], "overflow": f["overflow"],
			}
			for f in fields
		],
	}
	# Only added for club events, so a legacy event's etag is byte-for-byte what it
	# was before clubs existed - but two clubs sharing a slug get distinct etags,
	# so one can never be served the other's cached render.
	if club_slug:
		fingerprint_data["club"] = club_slug
	fingerprint = json.dumps(fingerprint_data, sort_keys=True, separators=(",", ":"))
	etag = hashlib.sha256(fingerprint.encode("utf-8")).hexdigest()

	# Checked before the semaphore: a cache hit costs microseconds and must never
	# queue behind someone else's 400 ms encode.
	cached = _cache_get(_RENDERED_CERT_CACHE, etag)
	if cached is not None:
		return cached

	with render_slot(tenant_key or club_slug or slug):
		# Re-checked inside the slot: whoever we queued behind may have been
		# rendering this very certificate.
		cached = _cache_get(_RENDERED_CERT_CACHE, etag)
		if cached is not None:
			return cached

		image = get_template_image(slug, config, club_slug)
		if image is None:
			return None

		draw_fields_on_image(image, fields)

		image_bytes, mimetype = encode_certificate(image, variant, download_format)

	result = (image_bytes, etag, mimetype)
	_cache_put(_RENDERED_CERT_CACHE, etag, result, _RENDER_CACHE_MAX)
	return result


# ─── Certificate links ────────────────────────────────────────────────────────
#
# A certificate link is a signed token carrying the event slug and the printed
# name, not a pointer to a stored file. Links are unguessable without SECRET_KEY,
# survive restarts and redeploys, and work on any worker.

_CERT_TOKEN_SALT = "certificate-link-v1"


def _cert_serializer() -> URLSafeSerializer:
	return URLSafeSerializer(app.secret_key, salt=_CERT_TOKEN_SALT)


def make_cert_token(slug: str, cert_values, club_slug: str | None = None) -> str:
	"""Mint a link carrying the club, the slug, and every field's resolved value.

	The club is signed into the token, so the club is never taken from an unsigned
	URL segment: a token minted for one club cannot be replayed against another.
	`cert_values` is a {field_id: text} map, or a bare string (stored as {"name": …}).
	"""
	values = {"name": cert_values} if isinstance(cert_values, str) else dict(cert_values or {})
	values = {str(k): str(v) for k, v in list(values.items())[:MAX_FIELDS]}
	payload = {"s": slug, "v": values}
	if club_slug:
		payload["c"] = club_slug
	return _cert_serializer().dumps(payload)


def read_cert_token(token: str) -> tuple[str | None, str, dict] | None:
	"""Return (club_slug_or_None, slug, {field_id: text}) or None.

	Tokens carry `c` (club), `s` (slug), `v` (values). A token with no `c` is a
	pre-Phase-3 legacy link; the caller decides whether the grace period still
	honours it. Pre-0.5 tokens carry `n` (a single name) instead of `v`.
	"""
	try:
		payload = _cert_serializer().loads(token)
	except BadSignature:
		return None
	if not isinstance(payload, dict):
		return None
	slug = str(payload.get("s", ""))
	if not safe_slug(slug):
		return None
	club = payload.get("c")
	club_slug = str(club) if club else None
	if club_slug is not None and not _safe_club_slug(club_slug):
		return None
	raw_values = payload.get("v")
	if isinstance(raw_values, dict):
		values = {str(k): str(v) for k, v in raw_values.items()}
	else:
		name = str(payload.get("n", ""))
		values = {"name": name} if name else {}
	if not any(v for v in values.values()):
		return None
	return club_slug, slug, values


def _legacy_cert_record(cert_id: str) -> tuple[str | None, str, dict] | None:
	"""Resolve a pre-token certificate id from its on-disk metadata file."""
	path = _cert_metadata_path(cert_id)
	if not os.path.exists(path):
		return None
	try:
		with open(path, encoding="utf-8") as f:
			metadata = json.load(f)
	except (OSError, json.JSONDecodeError):
		return None
	if not isinstance(metadata, dict):
		return None
	slug = str(metadata.get("event_slug", ""))
	cert_name = str(metadata.get("cert_name", ""))
	if not safe_slug(slug) or not cert_name:
		return None
	# Pre-token disk records predate clubs, so they are club-less (grace applies).
	return None, slug, {"name": cert_name}


def resolve_cert_token(token: str) -> tuple[str | None, str, dict] | None:
	"""Accept a signed token, or a legacy 32-hex id still sitting on local disk."""
	resolved = read_cert_token(token)
	if resolved is not None:
		return resolved
	if re.match(r"^[a-f0-9]{32}$", token):
		return _legacy_cert_record(token)
	return None


def draw_name_on_image(image: Image.Image, metadata: dict) -> None:
	draw = ImageDraw.Draw(image)
	font = get_font(metadata.get("font_size", 100), metadata.get("font_key"))
	raw_color = metadata.get("font_color", [50, 34, 24])
	try:
		color = tuple(int(channel) for channel in raw_color)[:3]
	except (TypeError, ValueError):
		color = (50, 34, 24)
	position = (metadata.get("text_x", 1789), metadata.get("text_y", 1440))
	text = metadata.get("cert_name", "")
	# anchor= is only supported by truetype fonts; load_default() is the fallback
	# when the bundled TTF is missing, and would raise instead of degrading.
	if isinstance(font, ImageFont.FreeTypeFont):
		draw.text(position, text, fill=color, font=font, anchor="mm")
	else:
		draw.text(position, text, fill=color, font=font)


def safe_download_name(name: str, slug: str, mimetype: str = "image/png") -> str:
	cleaned = re.sub(r"[^A-Za-z0-9 _-]", "", (name or "").strip())
	cleaned = re.sub(r"\s+", "-", cleaned)
	cleaned = cleaned.strip("-")
	if not cleaned:
		cleaned = f"{slug}-certificate"
	# The extension follows what was actually encoded, not what the event asked
	# for: an alpha template served as PNG must not be handed over named .jpg.
	return f"{cleaned}{DOWNLOAD_MIMETYPE_EXTENSIONS.get(mimetype, '.png')}"


def _has_image_magic(data: bytes, ext: str) -> bool:
	"""Check that the file content matches the extension it claims."""
	header = data[:12]
	if ext == ".png":
		return header[:8] == b"\x89PNG\r\n\x1a\n"
	if ext in (".jpg", ".jpeg"):
		return header[:3] == b"\xff\xd8\xff"
	if ext == ".gif":
		return header[:6] in (b"GIF87a", b"GIF89a")
	if ext == ".webp":
		return header[:4] == b"RIFF" and header[8:12] == b"WEBP"
	return False


def validate_template_upload(data: bytes, ext: str) -> str | None:
	"""Return an error message for a rejected template, or None when it is usable."""
	if not _has_image_magic(data, ext):
		return "File does not appear to be a valid image."
	try:
		with Image.open(BytesIO(data)) as probe:
			probe.verify()
		with Image.open(BytesIO(data)) as probe:
			width, height = probe.size
	except Exception:
		return "That image could not be decoded. Try re-exporting it."
	if width * height > MAX_TEMPLATE_PIXELS:
		megapixels = MAX_TEMPLATE_PIXELS // 1_000_000
		return f"Template is too large ({width}x{height}). The maximum is {megapixels} megapixels."
	return None


def _parse_int(value: str | None, fallback: int) -> int:
	try:
		return max(0, int(value)) if value is not None else fallback
	except (TypeError, ValueError):
		return fallback


def _parse_color(value: str | None, fallback: list | None = None) -> list[int]:
	if fallback is None:
		fallback = [50, 34, 24]
	if not value:
		return fallback
	value = value.strip().lstrip("#")
	if len(value) == 6:
		try:
			return [int(value[0:2], 16), int(value[2:4], 16), int(value[4:6], 16)]
		except ValueError:
			pass
	return fallback


def build_preview_metadata(event_config: dict, cert_name: str | None = None) -> dict:
	return {
		"event_slug": event_config.get("slug", ""),
		"cert_name": (cert_name or "Sample Text").strip() or "Sample Text",
		"text_x": _parse_int(request.args.get("text_x"), event_config.get("text_x", 1789)),
		"text_y": _parse_int(request.args.get("text_y"), event_config.get("text_y", 1440)),
		"font_size": _parse_int(request.args.get("font_size"), event_config.get("font_size", 100)),
		"font_color": _parse_color(request.args.get("font_color"), event_config.get("font_color", [50, 34, 24])),
		"font_key": normalize_font_key(request.args.get("font_key"), event_config.get("font_key", DEFAULT_FONT_KEY)),
	}


# A "Suggest values" dropdown publishes every distinct value of a column to the
# unauthenticated event page. For a team column that is a short, intended list;
# for a name, email, or roll-number column it is the entire roster. We cannot
# know a column's contents from its header alone, so this is a heuristic that
# decides how loudly to warn - never a silent block. The admin still chooses.
_PERSON_LEVEL_HINTS = (
	"name", "email", "mail", "phone", "mobile", "contact",
	"roll", "reg", "registration", "id", "player", "student",
	"participant", "member", "user", "person", "usn", "prn",
)


def looks_person_level(column: str) -> bool:
	"""True when a column header suggests per-person data that would leak a roster."""
	token = re.sub(r"[^a-z0-9]+", " ", (column or "").lower())
	words = token.split()
	for hint in _PERSON_LEVEL_HINTS:
		# Whole-word match, so "team" does not trip on "eam" and "id" does not
		# fire inside "video". A trailing "team_id" still matches on the "id" word.
		if hint in words:
			return True
	return False


@app.context_processor
def inject_style_context() -> dict:
	return {
		"font_options": available_font_options(),
		"default_font_key": normalize_font_key(DEFAULT_FONT_KEY),
		# A callable, so only templates that actually render a form mint a token.
		"csrf_token": csrf_token,
		"validation_type_labels": VALIDATION_TYPE_LABELS,
		"looks_person_level": looks_person_level,
	}


# ─── CSRF ─────────────────────────────────────────────────────────────────────
#
# Every state-changing request must carry a token tied to the caller's session.
# Enforcement is fail-closed and applies to all unsafe methods by default, so a
# route added later is protected without anyone remembering to protect it.

CSRF_FIELD_NAME = "csrf_token"
CSRF_HEADER_NAME = "X-CSRF-Token"
CSRF_UNSAFE_METHODS = ("POST", "PUT", "PATCH", "DELETE")
_CSRF_EXEMPT_ENDPOINTS: set[str] = set()


def csrf_exempt(view):
	"""Opt a route out of CSRF validation. Nothing needs this today."""
	_CSRF_EXEMPT_ENDPOINTS.add(view.__name__)
	return view


def csrf_token() -> str:
	"""Current session token, minted on first use. Exposed to templates."""
	token = session.get(CSRF_FIELD_NAME)
	if not isinstance(token, str) or not token:
		token = secrets.token_urlsafe(32)
		session[CSRF_FIELD_NAME] = token
	return token


def _submitted_csrf_token() -> str:
	form_value = request.form.get(CSRF_FIELD_NAME, "")
	if form_value:
		return form_value
	return request.headers.get(CSRF_HEADER_NAME, "")


@app.before_request
def _enforce_csrf():
	if request.method not in CSRF_UNSAFE_METHODS:
		return None
	if request.endpoint in _CSRF_EXEMPT_ENDPOINTS:
		return None

	expected = session.get(CSRF_FIELD_NAME, "")
	submitted = _submitted_csrf_token()
	if expected and submitted and hmac.compare_digest(str(expected), submitted):
		return None

	if request.headers.get("X-Requested-With") == "XMLHttpRequest":
		return jsonify({"ok": False, "error": "Session expired. Reload the page and try again."}), 400
	return render_template("csrf_error.html"), 400


# ─── Admin auth ───────────────────────────────────────────────────────────────

def require_admin(f):
	@wraps(f)
	def decorated(*args, **kwargs):
		if not session.get("admin_logged_in"):
			return redirect(url_for("admin_login"))
		return f(*args, **kwargs)
	return decorated


# The existing single-password gate is the SUPERADMIN surface now: it approves
# clubs and (through Phase 5) still manages the legacy csi-aseb events. A club
# session (club_id) can never reach it, and it can never reach a club dashboard -
# the two are keyed on different session fields, never shared.
require_superadmin = require_admin


def require_club(f):
	"""Gate a club-scoped route. A pending club is allowed in to configure; only a
	suspended or vanished club is bounced. Sets g.club for the handler."""
	@wraps(f)
	def decorated(*args, **kwargs):
		club_id = session.get("club_id")
		if not club_id:
			return redirect(url_for("club_login"))
		club = repo().get_club_by_id(club_id)
		if club is None or club["status"] == "suspended":
			session.pop("club_id", None)
			return redirect(url_for("club_login"))
		g.club = club
		return f(*args, **kwargs)
	return decorated


# ─── Public routes ────────────────────────────────────────────────────────────

@app.route("/", methods=["GET"])
def home():
	# Reflect an existing session so a signed-in club/admin gets a straight path to
	# their dashboard. The landing page is public, so a database hiccup degrades to
	# the logged-out view rather than 503-ing the whole page.
	club = None
	club_id = session.get("club_id")
	if club_id:
		try:
			club = repo().get_club_by_id(club_id)
		except db.DatabaseUnavailable:
			club = None
		if club is not None and club["status"] == "suspended":
			club = None
	return render_template("index.html", club=club,
						   admin_logged_in=bool(session.get("admin_logged_in")))


@app.route("/events/<slug>", methods=["GET"])
def event_page(slug: str):
	# Live links to the old flat URL survive: redirect to the club-scoped URL, but
	# only when csi-aseb actually has this event. A 301 to a page that 404s would
	# be cached by the browser and become very hard to undo.
	if not safe_slug(slug):
		abort(404)
	if resolve_public_event(LEGACY_TOKEN_CLUB, slug, require_active=True) is None:
		abort(404)
	return redirect(url_for("club_event_page", club_slug=LEGACY_TOKEN_CLUB, event_slug=slug), code=301)


@app.route("/events/<slug>/download", methods=["POST"])
def download_certificate(slug: str):
	if not safe_slug(slug):
		return redirect(url_for("home"))
	config = load_event(slug)
	if config is None or not config.get("active", False):
		return redirect(url_for("home"))
	validation_type = config.get("validation_type", "player_team")
	primary_id, _ = participant_input_fields(config)
	cert_name = (request.form.get("cert_name", "") or "").strip()
	# cert_name is only needed when a field is filled by participant input; a
	# csv-sourced name is read from the matched row, so it is not asked for.
	if primary_id and not cert_name:
		return render_template("event.html", **event_form_context(config, slug, "Please fill all fields.")), 400
	matched_rows, validation_error = validate_participant_submission(slug, config, request.form)
	if validation_error:
		return render_template("event.html", **event_form_context(config, slug, validation_error)), 400
	if not has_template(slug, config):
		return render_template("event.html", **event_form_context(config, slug, "Certificate template not found on server.")), 500

	# Resolve every field to its printed value. Input fields come from the form -
	# the primary name from `cert_name`, any additional input field from field_<id>.
	# csv and static fields resolve server-side and are never taken from the form.
	fields = normalize_fields(config)
	inputs = {}
	for field in fields:
		if field["source"] == "input":
			posted = request.form.get("field_" + field["id"])
			inputs[field["id"]] = (posted or "").strip()
	# The primary name box (cert_name) fills the primary input field unless that
	# field already got a dedicated field_<id> value.
	if primary_id and not inputs.get(primary_id):
		inputs[primary_id] = cert_name
	values, resolve_error = resolve_field_values(fields, matched_rows, inputs)
	if resolve_error:
		return render_template("event.html", **event_form_context(config, slug, resolve_error)), 400

	# The image is rendered on demand from the token, so this POST stays cheap.
	# The legacy flat route serves csi-aseb, so its token names that club.
	return redirect(url_for("preview_page", token=make_cert_token(slug, values, club_slug=LEGACY_TOKEN_CLUB)))


@app.route("/c/<club_slug>/<event_slug>", methods=["GET"])
def club_event_page(club_slug: str, event_slug: str):
	resolved = resolve_public_event(club_slug, event_slug, require_active=True)
	if resolved is None:
		abort(404)
	config, storage_club = resolved
	action = url_for("club_download_certificate", club_slug=club_slug, event_slug=event_slug)
	return render_template("event.html",
						   **event_form_context(config, event_slug, None, download_action=action,
												club_slug=storage_club))


@app.route("/c/<club_slug>/<event_slug>/download", methods=["POST"])
def club_download_certificate(club_slug: str, event_slug: str):
	resolved = resolve_public_event(club_slug, event_slug, require_active=True)
	if resolved is None:
		abort(404)
	config, storage_club = resolved
	action = url_for("club_download_certificate", club_slug=club_slug, event_slug=event_slug)

	def form_error(message, status=400):
		return render_template("event.html",
							   **event_form_context(config, event_slug, message, download_action=action,
													club_slug=storage_club)), status

	primary_id, _ = participant_input_fields(config)
	cert_name = (request.form.get("cert_name", "") or "").strip()
	if primary_id and not cert_name:
		return form_error("Please fill all fields.")
	matched_rows, validation_error = validate_participant_submission(event_slug, config, request.form, club_slug=storage_club)
	if validation_error:
		return form_error(validation_error)
	if not has_template(event_slug, config):
		return form_error("Certificate template not found on server.", 500)

	fields = normalize_fields(config)
	inputs = {}
	for field in fields:
		if field["source"] == "input":
			posted = request.form.get("field_" + field["id"])
			inputs[field["id"]] = (posted or "").strip()
	if primary_id and not inputs.get(primary_id):
		inputs[primary_id] = cert_name
	values, resolve_error = resolve_field_values(fields, matched_rows, inputs)
	if resolve_error:
		return form_error(resolve_error)

	# The club is signed into the token, so the club can never be spoofed by the
	# URL at render time - the token minted here resolves to this club and no other.
	token = make_cert_token(event_slug, values, club_slug=club_slug)
	return redirect(url_for("preview_page", token=token))


@app.route("/c/<club_slug>", methods=["GET"])
def club_public(club_slug: str):
	# A club's public page is dark unless approved. 404 - never 403 - so a probe
	# cannot tell an unapproved club from one that does not exist.
	if not _safe_club_slug(club_slug):
		abort(404)
	club = repo().get_club_by_slug(club_slug)
	if club is None:
		# csi-aseb has no club row until Phase 5: list its active legacy events.
		if club_slug == LEGACY_TOKEN_CLUB:
			legacy = [{"name": e.get("name", e.get("slug", "")), "slug": e.get("slug", "")}
					  for e in load_all_events(active_only=True)]
			return render_template("club/public.html",
								   club={"name": LEGACY_TOKEN_CLUB, "slug": LEGACY_TOKEN_CLUB},
								   events=legacy, club_slug=club_slug)
		abort(404)
	if club["status"] != "approved":
		abort(404)
	events = [e for e in repo().list_events(club["id"]) if e.get("active")]
	return render_template("club/public.html", club=club, events=events, club_slug=club_slug)


def _log_legacy_token(slug: str) -> None:
	"""Record every club-less token resolution. When this goes quiet the grace
	period is over and the legacy branch can be removed."""
	log_event("legacy-token", f"resolved club-less certificate token for slug='{slug}'")


# Club-event resolution is on the hot render path: _certificate_from_token calls
# it for every preview and download. Without a cache, each call is two Postgres
# round-trips (get_club_by_slug + get_event), so a 1000-participant spike would be
# thousands of pooler queries for a workload the render cache serves from memory.
# Cached per (club, event) with the same short TTL as the legacy config cache, and
# invalidated on any club-event mutation so publish/unpublish reflect immediately.
_PUBLIC_EVENT_CACHE: dict[tuple[str, str], tuple[dict, str | None, float]] = {}
_PUBLIC_EVENT_CACHE_TTL_SEC = 30.0


def _invalidate_public_event(club_slug: str, event_slug: str) -> None:
	_PUBLIC_EVENT_CACHE.pop((club_slug, event_slug), None)


def _invalidate_club_public_events(club_slug: str) -> None:
	"""Drop every cached public event for a club (e.g. on suspend/approve)."""
	for key in [k for k in _PUBLIC_EVENT_CACHE if k[0] == club_slug]:
		_PUBLIC_EVENT_CACHE.pop(key, None)


def _resolve_public_event_source(club_slug: str, event_slug: str) -> tuple[dict, str | None] | None:
	"""The (config, storage_club) for a public event, ignoring the active flag.

	The real-club Postgres path is cached; the csi-aseb bridge path is not, because
	load_event already caches it. Returns the SHARED cached dict - the caller must
	copy before handing it out or applying per-request state.
	"""
	key = (club_slug, event_slug)
	cached = _PUBLIC_EVENT_CACHE.get(key)
	if cached is not None and (time.time() - cached[2]) < _PUBLIC_EVENT_CACHE_TTL_SEC:
		return cached[0], cached[1]

	# The csi-aseb bridge: there is no club row for csi-aseb until Phase 5, so its
	# events resolve from the legacy KV/file store. A database outage must not take
	# this legacy serving down, so it degrades rather than 503-ing.
	try:
		club = repo().get_club_by_slug(club_slug)
	except db.DatabaseUnavailable:
		if club_slug == LEGACY_TOKEN_CLUB:
			club = None
		else:
			raise

	if club is None:
		if club_slug != LEGACY_TOKEN_CLUB:
			return None
		config = load_event(event_slug)  # already cached via _EVENT_CONFIG_CACHE
		if config is None:
			return None
		return config, None

	if club["status"] != "approved":
		return None
	event = repo().get_event(club["id"], event_slug)
	if event is None:
		return None
	config = dict(event.get("config") or {})
	config["slug"] = event_slug
	config.setdefault("name", event.get("name", ""))
	config["active"] = bool(event.get("active"))
	_PUBLIC_EVENT_CACHE[key] = (config, club["slug"], time.time())
	if len(_PUBLIC_EVENT_CACHE) > 500:
		_PUBLIC_EVENT_CACHE.clear()
		_PUBLIC_EVENT_CACHE[key] = (config, club["slug"], time.time())
	return config, club["slug"]


def resolve_public_event(club_slug: str, event_slug: str,
						 require_active: bool = True) -> tuple[dict, str | None] | None:
	"""Resolve a public event to (config, storage_club_slug), or None (-> 404).

	`storage_club_slug` is the object-store prefix: the club's slug for a real club
	event, or None for a legacy csi-aseb event that still lives at the bare path
	(until Phase 5). Isolation is by (club_id, slug), never slug alone, so club B
	asking for club A's event slug simply gets None -> 404, never 403.
	"""
	if not _safe_club_slug(club_slug) or not safe_slug(event_slug):
		return None
	resolved = _resolve_public_event_source(club_slug, event_slug)
	if resolved is None:
		return None
	config, storage_club = resolved
	if require_active and not config.get("active", False):
		return None
	# Copy: the source may be the shared cache entry, and callers mutate what they get.
	return copy.deepcopy(config), storage_club


def _certificate_from_token(token: str) -> tuple[str, dict, dict, str | None] | None:
	"""Resolve a certificate link to (slug, values, config, storage_club_slug).

	The club comes from the SIGNED token, never from the URL. A club-less legacy
	token resolves to csi-aseb only while the grace period is on, and every such
	resolution is logged. A deactivated event's already-issued link still renders
	(links are stateless and survive deactivation), but a token for an unapproved
	or vanished club resolves to nothing.
	"""
	resolved = resolve_cert_token(token)
	if resolved is None:
		return None
	club_slug, slug, values = resolved
	if club_slug is None:
		if not LEGACY_TOKEN_GRACE:
			return None
		_log_legacy_token(slug)
		club_slug = LEGACY_TOKEN_CLUB
	resolved_event = resolve_public_event(club_slug, slug, require_active=False)
	if resolved_event is None:
		return None
	config, storage_club = resolved_event
	return slug, values, config, storage_club


@app.route("/preview/<token>", methods=["GET"])
def preview_page(token: str):
	resolved = _certificate_from_token(token)
	if resolved is None:
		return redirect(url_for("home"))
	_, values, config, _club = resolved
	return render_template("preview.html", cert_token=token,
						   cert_name=display_name(values), event=config)


def render_busy_response():
	"""
	503 with Retry-After, for a render that could not get a slot.

	Shedding load explicitly beats holding the connection until gunicorn's
	--timeout kills the worker, which would drop every other request it was
	serving too.
	"""
	response = make_response("Certificate rendering is busy. Please retry in a moment.", 503)
	response.headers["Retry-After"] = "5"
	response.cache_control.no_store = True
	return response


@app.route("/preview-image/<token>", methods=["GET"])
def preview_image(token: str):
	resolved = _certificate_from_token(token)
	if resolved is None:
		return ("Not found", 404)
	slug, values, config, storage_club = resolved
	try:
		rendered = render_certificate(slug, values, config, variant="preview", club_slug=storage_club)
	except RenderCapacityError:
		return render_busy_response()
	if rendered is None:
		return ("Not found", 404)
	image_bytes, etag, mimetype = rendered

	response = send_file(BytesIO(image_bytes), mimetype=mimetype, etag=etag)
	# The image carries a participant's name, and it changes whenever an admin
	# edits the template or coordinates, so: private, revalidated, never immutable.
	response.cache_control.private = True
	response.cache_control.max_age = 300
	response.cache_control.must_revalidate = True
	return response


@app.route("/download-file/<token>", methods=["GET"])
def download_file(token: str):
	resolved = _certificate_from_token(token)
	if resolved is None:
		return ("Not found", 404)
	slug, values, config, storage_club = resolved
	try:
		rendered = render_certificate(slug, values, config, variant="download", club_slug=storage_club)
	except RenderCapacityError:
		return render_busy_response()
	if rendered is None:
		return ("Not found", 404)
	image_bytes, _, mimetype = rendered

	return send_file(
		BytesIO(image_bytes),
		mimetype=mimetype,
		as_attachment=True,
		download_name=safe_download_name(display_name(values), slug, mimetype),
	)


# ─── Admin routes ─────────────────────────────────────────────────────────────

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
	error = None
	if request.method == "POST":
		password = request.form.get("password", "")
		if not ADMIN_PASSWORD:
			error = "ADMIN_PASSWORD environment variable is not set on this server."
		elif hmac.compare_digest(password, ADMIN_PASSWORD):
			# New token for the newly privileged session, so a token an attacker
			# may have observed pre-login cannot be replayed against admin routes.
			session.clear()
			session.permanent = True
			session["admin_logged_in"] = True
			csrf_token()
			return redirect(url_for("admin_clubs"))
		else:
			error = "Incorrect password."
	return render_template("admin/login.html", error=error)


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
	session.clear()
	return redirect(url_for("admin_login"))


# The legacy single-tenant event manager. Kept at an obscure path for testing and
# for the migrated csi-aseb events; the superadmin home is the clubs dashboard now.
@app.route("/admin/legacy-events", methods=["GET"])
@require_admin
def admin_dashboard():
	return render_template("admin/dashboard.html", events=load_all_events())


# ─── Club auth & dashboard ────────────────────────────────────────────────────
#
# Self-registration creates a pending club. A pending club can log in and reach
# its dashboard to configure; its public pages stay dark (404) until a superadmin
# approves it. Passwords are stored only as werkzeug hashes.

CLUB_PASSWORD_MIN = 8


def _safe_club_slug(slug: str) -> bool:
	return bool(_SLUG_RE.match(slug)) and ".." not in slug and 3 <= len(slug) <= 40


@app.route("/register", methods=["GET", "POST"])
def club_register():
	if request.method == "GET":
		return render_template("auth/register.html", error=None, form={})
	name = (request.form.get("name", "") or "").strip()
	slug = (request.form.get("slug", "") or "").strip().lower()
	password = request.form.get("password", "")
	form = {"name": name, "slug": slug}

	def fail(message):
		return render_template("auth/register.html", error=message, form=form), 400

	if not name or not slug or not password:
		return fail("Club name, address, and password are all required.")
	if not _safe_club_slug(slug):
		return fail("Address must be 3-40 characters, lowercase letters, numbers, and hyphens.")
	if len(password) < CLUB_PASSWORD_MIN:
		return fail(f"Password must be at least {CLUB_PASSWORD_MIN} characters.")
	if repo().get_club_by_slug(slug) is not None:
		return fail("That address is already taken. Choose another.")

	try:
		club = repo().create_club(slug, name, generate_password_hash(password))
	except ValueError:
		return fail("That address is already taken. Choose another.")

	# A fresh session for the newly created club, so a pre-registration token
	# cannot be replayed against the dashboard.
	session.clear()
	session.permanent = True
	session["club_id"] = club["id"]
	csrf_token()
	return redirect(url_for("club_dashboard"))


@app.route("/login", methods=["GET", "POST"])
def club_login():
	if request.method == "GET":
		return render_template("auth/login.html", error=None)
	slug = (request.form.get("slug", "") or "").strip().lower()
	password = request.form.get("password", "")
	club = repo().get_club_by_slug(slug) if slug else None
	# One generic message whether the club is unknown or the password is wrong,
	# and check_password_hash always runs on a real-looking hash so timing does
	# not distinguish "no such club" from "bad password".
	reference = club["password_hash"] if club else generate_password_hash("_")
	# Computed once: the KDF is deliberately slow, so run it a single time and then
	# branch on status. Still runs on a real-looking hash when the club is unknown,
	# so timing does not distinguish "no such club" from "bad password".
	credentials_ok = club is not None and check_password_hash(reference, password)
	if credentials_ok and club["status"] != "suspended":
		session.clear()
		session.permanent = True
		session["club_id"] = club["id"]
		csrf_token()
		return redirect(url_for("club_dashboard"))
	if credentials_ok and club["status"] == "suspended":
		return render_template("auth/login.html",
							   error="This club has been suspended. Contact the administrator."), 403
	return render_template("auth/login.html", error="Incorrect club address or password."), 400


@app.route("/logout", methods=["POST"])
def club_logout():
	session.pop("club_id", None)
	return redirect(url_for("club_login"))


@app.route("/clubs/autocomplete", methods=["GET"])
def club_autocomplete():
	# Returns nothing under three characters, so the full club list is never
	# handed out on page load.
	prefix = (request.args.get("q", "") or "").strip().lower()
	if len(prefix) < 3:
		return jsonify({"slugs": []})
	return jsonify({"slugs": repo().search_club_slugs(prefix)})


@app.route("/dashboard", methods=["GET"])
@require_club
def club_dashboard():
	club = g.club
	events = repo().list_events(club["id"])
	used = club_storage_bytes(club["slug"])
	quota = int(club["quota_bytes"])
	return render_template("club/dashboard.html", club=club, events=events,
						   pending=(club["status"] == "pending"),
						   active_count=sum(1 for e in events if e.get("active")),
						   used_mb=used // (1024 * 1024), quota_mb=quota // (1024 * 1024),
						   used_pct=min(100, round(used * 100 / quota)) if quota else 0)


def _event_detail(club, slug, error=None, success=None, status=200):
	event = repo().get_event(club["id"], slug)
	if event is None:
		abort(404)
	config = event.get("config") or {}
	used = club_storage_bytes(club["slug"])
	return render_template(
		"club/event_detail.html", club=club, event=event,
		has_template=bool(config.get("template_version")),
		used_mb=used // (1024 * 1024), quota_mb=int(club["quota_bytes"]) // (1024 * 1024),
		error=error, success=success), status


@app.route("/dashboard/events", methods=["POST"])
@require_club
def club_create_event():
	club = g.club
	name = (request.form.get("name", "") or "").strip()
	slug = (request.form.get("slug", "") or "").strip().lower()
	events = repo().list_events(club["id"])
	if not name or not slug or not safe_slug(slug):
		return render_template("club/dashboard.html", club=club, events=events,
							   pending=(club["status"] == "pending"),
							   error="Give the event a name and a valid address (lowercase letters, numbers, hyphens)."), 400
	try:
		created = repo().create_event(club["id"], slug, name, {})
	except ValueError:
		return render_template("club/dashboard.html", club=club, events=events,
							   pending=(club["status"] == "pending"),
							   error=f"You already have an event at '{slug}'."), 400
	if created is None:
		abort(404)
	return redirect(url_for("club_event_detail", slug=slug))


@app.route("/dashboard/events/<slug>", methods=["GET"])
@require_club
def club_event_detail(slug: str):
	if not safe_slug(slug):
		abort(404)
	return _event_detail(g.club, slug)


@app.route("/dashboard/events/<slug>/template", methods=["POST"])
@require_club
def club_upload_template(slug: str):
	club = g.club
	if not safe_slug(slug):
		abort(404)
	event = repo().get_event(club["id"], slug)
	if event is None:
		abort(404)

	file = request.files.get("template_file")
	if not file or file.filename == "":
		return _event_detail(club, slug, error="No file selected.", status=400)
	ext = os.path.splitext(secure_filename(file.filename).lower())[1].lower()
	if ext not in TEMPLATE_EXTENSIONS:
		return _event_detail(club, slug, error="Template must be PNG, JPG, GIF, or WebP.", status=400)

	file.stream.seek(0)
	data = file.stream.read()
	# Magic-byte and decompression-bomb validation - the extension is never trusted.
	validation_error = validate_template_upload(data, ext)
	if validation_error:
		return _event_detail(club, slug, error=validation_error, status=400)

	# Quota is checked before anything is written, and a rejection stores nothing.
	# A replacement upserts over the existing template, so net that object out.
	old_ext = (event.get("config") or {}).get("template_ext")
	replacing = _template_object_path(slug, old_ext, club["slug"]) if old_ext else None
	ok, message = quota_status(club, len(data), replacing_path=replacing)
	if not ok:
		return _event_detail(club, slug, error=message, status=400)

	try:
		version = save_template_bytes(slug, data, ext, club_slug=club["slug"])
	except Exception:
		return _event_detail(club, slug, error="Could not save the template to storage. Try again.", status=502)

	config = dict(event.get("config") or {})
	config["template_ext"] = ext
	config["template_version"] = version
	repo().update_event(club["id"], slug, config=config, template_ext=ext, template_version=version)
	_invalidate_public_event(club["slug"], slug)
	warm_template_cache_from_bytes(slug, version, data, club_slug=club["slug"])
	return _event_detail(club, slug, success="Template uploaded.")


@app.route("/dashboard/events/<slug>/csv", methods=["POST"])
@require_club
def club_upload_csv(slug: str):
	club = g.club
	if not safe_slug(slug):
		abort(404)
	event = repo().get_event(club["id"], slug)
	if event is None:
		abort(404)

	file = request.files.get("csv_file")
	if not file or file.filename == "":
		return _event_detail(club, slug, error="No file selected.", status=400)
	ext = os.path.splitext(secure_filename(file.filename).lower())[1]
	if ext not in PARTICIPANT_EXTENSIONS:
		return _event_detail(club, slug, error="Participants file must be a .csv or .xlsx.", status=400)

	raw_upload = file.stream.read()
	content, upload_error = participant_text_from_upload(raw_upload, ext)
	if upload_error:
		return _event_detail(club, slug, error=upload_error, status=400)

	# A re-upload replaces the same data.csv, so net the existing object out.
	replacing = _participants_object_path(slug, "data.csv", club_slug=club["slug"])
	ok, message = quota_status(club, len(content.encode("utf-8")), replacing_path=replacing)
	if not ok:
		return _event_detail(club, slug, error=message, status=400)

	source = (raw_upload, ext) if ext != ".csv" else (content.encode("utf-8"), ".csv")
	try:
		save_event_csv(slug, content, source=source, club_slug=club["slug"])
	except Exception:
		return _event_detail(club, slug, error="Could not save the participant list. Try again.", status=502)
	_invalidate_public_event(club["slug"], slug)
	return _event_detail(club, slug, success="Participant list uploaded.")


def _club_event_or_404(club, slug):
	if not safe_slug(slug):
		abort(404)
	event = repo().get_event(club["id"], slug)
	if event is None:
		abort(404)
	return event


@app.route("/dashboard/events/<slug>/config", methods=["POST"])
@require_club
def club_update_config(slug):
	club = g.club
	event = _club_event_or_404(club, slug)
	config = dict(event.get("config") or {})
	name = (request.form.get("name", "") or event.get("name") or "").strip()
	validation_type = request.form.get("validation_type", config.get("validation_type", "player_team"))
	if validation_type not in VALIDATION_TYPES:
		validation_type = "player_team"
	config["validation_type"] = validation_type
	# Only the "custom" type carries custom fields, and only then does the settings
	# form submit them. For every other type, leave any existing custom config
	# untouched rather than wiping it (the form has no field to resubmit it).
	if validation_type == "custom":
		parsed = parse_custom_fields(request.form.getlist("custom_fields"))
		config["custom_fields"] = parsed
		existing_drop = config.get("custom_dropdown_fields") or []
		config["custom_dropdown_fields"] = [f for f in existing_drop if f in parsed]
	config["download_format"] = normalize_download_format(
		request.form.get("download_format"), config.get("download_format", DEFAULT_DOWNLOAD_FORMAT))
	repo().update_event(club["id"], slug, name=name, config=config)
	_invalidate_public_event(club["slug"], slug)
	return _event_detail(club, slug, success="Settings saved.")


@app.route("/dashboard/events/<slug>/toggle", methods=["POST"])
@require_club
def club_toggle_event(slug):
	club = g.club
	event = _club_event_or_404(club, slug)
	# Publishing needs a template, or the public page would 500 on the first visit.
	going_active = not bool(event.get("active"))
	if going_active and not (event.get("config") or {}).get("template_version"):
		return _event_detail(club, slug, error="Upload a template before publishing this event.", status=400)
	repo().update_event(club["id"], slug, active=going_active)
	_invalidate_public_event(club["slug"], slug)
	if request.form.get("next") == "dashboard":
		return redirect(url_for("club_dashboard"))
	return redirect(url_for("club_event_detail", slug=slug))


@app.route("/dashboard/events/<slug>/delete", methods=["POST"])
@require_club
def club_delete_event(slug):
	club = g.club
	event = _club_event_or_404(club, slug)
	# Typed-confirmation, matching the superadmin delete: the slug must be retyped.
	if request.form.get("confirm", "") != slug:
		return _event_detail(club, slug, error="Type the event address to confirm deletion.", status=400)
	delete_event_storage(slug, club_slug=club["slug"])
	repo().delete_event(club["id"], slug)
	_invalidate_public_event(club["slug"], slug)
	return redirect(url_for("club_dashboard"))


@app.route("/dashboard/events/<slug>/template-preview", methods=["GET"])
@require_club
def club_template_preview(slug):
	club = g.club
	event = _club_event_or_404(club, slug)
	config = event.get("config") or {}
	data = load_template_bytes(slug, config, club["slug"])
	if data is None:
		return "Template not found", 404
	ext = template_ext_for(slug, config)
	response = send_file(BytesIO(data),
						 mimetype=TEMPLATE_CONTENT_TYPES.get(ext, "application/octet-stream"),
						 etag=(config.get("template_version") or "none"))
	response.cache_control.private = True
	response.cache_control.max_age = 60
	return response


@app.route("/dashboard/events/<slug>/coordinates", methods=["GET"])
@require_club
def club_coordinate_editor(slug):
	club = g.club
	event = _club_event_or_404(club, slug)
	config = event.get("config") or {}
	config.setdefault("slug", slug)
	config.setdefault("name", event.get("name", slug))
	dataset = load_participant_dataset(slug, club["slug"])
	sample_rows = dataset.raw_rows[:25]
	fonts = [
		{"key": o["key"], "label": o["label"], "family": o["css_family"],
		 "weight": o["css_weight"], "url": url_for("font_asset", font_key=o["key"])}
		for o in available_font_options()
	]
	return render_template(
		"admin/coordinate_editor.html",
		event=config,
		fields_json=_json_for_script(normalize_fields(config)),
		columns_json=_json_for_script(dataset.columns),
		sample_rows_json=_json_for_script(sample_rows),
		fonts_json=_json_for_script(fonts),
		max_fields=MAX_FIELDS,
		has_csv=bool(dataset.raw_rows),
		save_url=url_for("club_save_fields", slug=slug),
		template_url=url_for("club_template_preview", slug=slug),
		back_url=url_for("club_event_detail", slug=slug),
	)


@app.route("/dashboard/events/<slug>/fields", methods=["POST"])
@require_club
def club_save_fields(slug):
	club = g.club
	if not safe_slug(slug):
		return jsonify({"ok": False, "error": "Unknown event."}), 404
	event = repo().get_event(club["id"], slug)
	if event is None:
		return jsonify({"ok": False, "error": "Unknown event."}), 404
	payload = request.get_json(silent=True) or {}
	fields, error = validate_fields_payload(slug, payload.get("fields"), club_slug=club["slug"])
	if error:
		return jsonify({"ok": False, "error": error}), 400
	config = dict(event.get("config") or {})
	config["fields"] = fields
	repo().update_event(club["id"], slug, config=config)
	return jsonify({"ok": True, "message": "Placement saved.", "fields": fields})

	return _event_detail(club, slug, success="Participant list uploaded.")




# ─── Superadmin: club approvals ───────────────────────────────────────────────

# The superadmin home: manage clubs. Also answers /admin/clubs so old links survive.
@app.route("/admin", methods=["GET"])
@app.route("/admin/clubs", methods=["GET"])
@require_superadmin
def admin_clubs():
	return render_template("admin/clubs.html", clubs=repo().list_clubs())


@app.route("/admin/clubs/<club_id>/status", methods=["POST"])
@require_superadmin
def admin_set_club_status(club_id: str):
	status = request.form.get("status", "")
	if status not in db.CLUB_STATUSES:
		return redirect(url_for("admin_clubs"))
	repo().set_club_status(club_id, status)
	# Suspending/approving flips a club's whole public surface, so drop its cached
	# public-event entries - a suspended club must go dark promptly, not after the TTL.
	club = repo().get_club_by_id(club_id)
	if club:
		_invalidate_club_public_events(club["slug"])
	return redirect(url_for("admin_clubs"))


@app.route("/admin/clubs/<club_id>/quota", methods=["POST"])
@require_superadmin
def admin_set_club_quota(club_id: str):
	try:
		quota_mb = max(1, int(request.form.get("quota_mb", "")))
	except (TypeError, ValueError):
		return redirect(url_for("admin_clubs"))
	repo().set_club_quota(club_id, quota_mb * 1024 * 1024)
	return redirect(url_for("admin_clubs"))


@app.route("/admin/clubs/<club_id>/reset-password", methods=["POST"])
@require_superadmin
def admin_reset_club_password(club_id: str):
	new_password = request.form.get("password", "")
	if len(new_password) >= CLUB_PASSWORD_MIN:
		repo().set_club_password(club_id, generate_password_hash(new_password))
	return redirect(url_for("admin_clubs"))


@app.route("/admin/events/new", methods=["GET"])
@require_admin
def admin_new_event():
	return render_template("admin/event_form.html", event=None, is_new=True, error=None)


@app.route("/admin/events/new", methods=["POST"])
@require_admin
def admin_create_event():
	name = (request.form.get("name", "") or "").strip()
	slug = (request.form.get("slug", "") or "").strip().lower()
	validation_type = request.form.get("validation_type", "player_team")
	if validation_type not in VALIDATION_TYPES:
		validation_type = "player_team"
	custom_fields = parse_custom_fields(request.form.getlist("custom_fields"))
	custom_dropdown_fields = [field for field in parse_custom_fields(request.form.getlist("custom_dropdown_fields")) if field in custom_fields]
	text_x = _parse_int(request.form.get("text_x"), 1789)
	text_y = _parse_int(request.form.get("text_y"), 1440)
	font_size = _parse_int(request.form.get("font_size"), 100)
	font_color = _parse_color(request.form.get("font_color", ""))
	font_key = normalize_font_key(request.form.get("font_key"), DEFAULT_FONT_KEY)
	# Written explicitly on every new event, so an absent key only ever describes a
	# config that predates the format option and must keep rendering PNG.
	download_format = normalize_download_format(request.form.get("download_format"),
											   NEW_EVENT_DOWNLOAD_FORMAT)
	form_data = {"name": name, "slug": slug, "validation_type": validation_type, "custom_fields": custom_fields,
				 "custom_dropdown_fields": custom_dropdown_fields,
				 "text_x": text_x, "text_y": text_y, "font_size": font_size, "font_color": font_color,
				 "font_key": font_key, "download_format": download_format}
	if not name or not slug:
		return render_template("admin/event_form.html", event=form_data, is_new=True, error="Name and slug are required."), 400
	if not safe_slug(slug):
		return render_template("admin/event_form.html", event=form_data, is_new=True,
							   error="Slug must be lowercase letters, numbers, and hyphens only."), 400
	if _event_exists(slug):
		return render_template("admin/event_form.html", event=form_data, is_new=True,
							   error=f"An event with slug '{slug}' already exists."), 400
	os.makedirs(_event_dir(slug), exist_ok=True)
	config = {"name": name, "slug": slug, "active": False, "validation_type": validation_type, "custom_fields": custom_fields,
			  "custom_dropdown_fields": custom_dropdown_fields,
			  "text_x": text_x, "text_y": text_y, "font_size": font_size, "font_color": font_color, "font_key": font_key,
			  "download_format": download_format}
	save_event_config(slug, config)
	_set_event_state(slug, deleted=False, active=False)
	return redirect(url_for("admin_edit_event", slug=slug))


@app.route("/admin/events/<slug>", methods=["GET"])
@require_admin
def admin_edit_event(slug: str):
	if not safe_slug(slug):
		return redirect(url_for("admin_dashboard"))
	config = load_event(slug)
	if config is None:
		return redirect(url_for("admin_dashboard"))
	return render_template("admin/event_form.html", event=config, is_new=False, error=None,
						   has_template=has_template(slug, config),
						   has_csv=_event_csv_exists(slug),
						   csv_columns=csv_headers(slug))


@app.route("/admin/events/<slug>/config", methods=["POST"])
@require_admin
def admin_update_config(slug: str):
	if not safe_slug(slug):
		return redirect(url_for("admin_dashboard"))
	config = load_event(slug)
	if config is None:
		return redirect(url_for("admin_dashboard"))
	config["name"] = (request.form.get("name", "") or config["name"]).strip()
	validation_type = request.form.get("validation_type", config.get("validation_type", "player_team"))
	if validation_type not in VALIDATION_TYPES:
		validation_type = "player_team"
	config["validation_type"] = validation_type
	parsed_custom_fields = parse_custom_fields(request.form.getlist("custom_fields"))
	parsed_custom_dropdown_fields = parse_custom_fields(request.form.getlist("custom_dropdown_fields"))
	allowed_dropdown_fields = [field for field in parsed_custom_dropdown_fields if field in parsed_custom_fields]
	if validation_type == "custom":
		config["custom_fields"] = parsed_custom_fields or config.get("custom_fields", [])
	else:
		config["custom_fields"] = parsed_custom_fields
	config["custom_dropdown_fields"] = [field for field in allowed_dropdown_fields if field in config.get("custom_fields", [])]
	config["text_x"] = _parse_int(request.form.get("text_x"), config.get("text_x", 1789))
	config["text_y"] = _parse_int(request.form.get("text_y"), config.get("text_y", 1440))
	config["font_size"] = _parse_int(request.form.get("font_size"), config.get("font_size", 100))
	config["font_color"] = _parse_color(request.form.get("font_color"), config.get("font_color", [50, 34, 24]))
	config["font_key"] = normalize_font_key(request.form.get("font_key"), config.get("font_key", DEFAULT_FONT_KEY))
	# A legacy config with no key falls back to PNG and gets it written down, which
	# preserves its current output while making the choice visible from then on.
	config["download_format"] = normalize_download_format(
		request.form.get("download_format"),
		config.get("download_format", DEFAULT_DOWNLOAD_FORMAT))
	save_event_config(slug, config)
	if request.headers.get("X-Requested-With") == "XMLHttpRequest":
		return jsonify({"ok": True, "message": "Settings saved."})
	return render_template("admin/event_form.html", event=config, is_new=False, success="Settings saved.",
						   error=None, has_template=has_template(slug, config),
						   has_csv=_event_csv_exists(slug),
						   csv_columns=csv_headers(slug))


@app.route("/admin/events/<slug>/upload-template", methods=["POST"])
@require_admin
def admin_upload_template(slug: str):
	if not safe_slug(slug):
		return redirect(url_for("admin_dashboard"))
	config = load_event(slug)
	if config is None:
		return redirect(url_for("admin_dashboard"))
	has_csv = _event_csv_exists(slug)
	template_present = has_template(slug, config)
	file = request.files.get("template_file")
	if not file or file.filename == "":
		return render_template("admin/event_form.html", event=config, is_new=False,
							   error="No file selected.", has_template=template_present, has_csv=has_csv,
							   csv_columns=csv_headers(slug)), 400

	filename = secure_filename(file.filename).lower()
	ext = os.path.splitext(filename)[1].lower()
	if ext not in TEMPLATE_EXTENSIONS:
		return render_template("admin/event_form.html", event=config, is_new=False,
				error="Template must be PNG, JPG, GIF, or WebP.", has_template=template_present, has_csv=has_csv,
				csv_columns=csv_headers(slug)), 400

	file.stream.seek(0)
	data = file.stream.read()
	validation_error = validate_template_upload(data, ext)
	if validation_error:
		return render_template("admin/event_form.html", event=config, is_new=False,
				error=validation_error, has_template=template_present, has_csv=has_csv,
				csv_columns=csv_headers(slug)), 400

	try:
		version = save_template_bytes(slug, data, ext)
	except Exception:
		return render_template("admin/event_form.html", event=config, is_new=False,
				error="Could not save the template to storage. Check the Supabase settings and try again.",
				has_template=template_present, has_csv=has_csv,
				csv_columns=csv_headers(slug)), 502

	# Recording the version on the config is what invalidates every image cache:
	# render keys embed it, so a replaced template can never be served stale.
	config["template_ext"] = ext
	config["template_version"] = version
	save_event_config(slug, config)
	warm_template_cache_from_bytes(slug, version, data)
	return render_template("admin/event_form.html", event=config, is_new=False,
				success="Template uploaded successfully.", error=None, has_template=True, has_csv=has_csv,
				csv_columns=csv_headers(slug))


@app.route("/admin/events/<slug>/upload-csv", methods=["POST"])
@require_admin
def admin_upload_csv(slug: str):
	if not safe_slug(slug):
		return redirect(url_for("admin_dashboard"))
	config = load_event(slug)
	if config is None:
		return redirect(url_for("admin_dashboard"))
	template_present = has_template(slug, config)
	has_csv = _event_csv_exists(slug)
	file = request.files.get("csv_file")
	if not file or file.filename == "":
		return render_template("admin/event_form.html", event=config, is_new=False,
							   error="No file selected.", has_template=template_present, has_csv=has_csv,
							   csv_columns=csv_headers(slug)), 400
	filename = secure_filename(file.filename).lower()
	ext = os.path.splitext(filename)[1]
	if ext not in PARTICIPANT_EXTENSIONS:
		return render_template("admin/event_form.html", event=config, is_new=False,
							   error="Participants file must be a .csv or .xlsx.",
							   has_template=template_present, has_csv=has_csv,
							   csv_columns=csv_headers(slug)), 400

	raw_upload = file.stream.read()
	content, upload_error = participant_text_from_upload(raw_upload, ext)
	if upload_error:
		return render_template("admin/event_form.html", event=config, is_new=False,
							   error=upload_error, has_template=template_present, has_csv=has_csv,
							   csv_columns=csv_headers(slug)), 400

	validation_type = config.get("validation_type", "player_team")
	custom_fields: list[str] = config.get("custom_fields", [])
	if validation_type == "custom" and not custom_fields:
		return render_template("admin/event_form.html", event=config, is_new=False,
							   error="Pick at least one column to match on before uploading a participant list.",
							   has_template=template_present, has_csv=has_csv,
							   csv_columns=csv_headers(slug)), 400
	required_headers = required_headers_for_validation(validation_type, custom_fields)
	try:
		reader = csv.DictReader(content.splitlines())
		headers = {h.strip().lower() for h in (reader.fieldnames or [])}
		if validation_type == "badge_id" and not ({"roll_no", "id", "badge_id", "badge_number"} & headers):
			return render_template("admin/event_form.html", event=config, is_new=False,
								   error="The file must include one of: roll_no, id, badge_id, badge_number.",
								   has_template=template_present, has_csv=has_csv,
								   csv_columns=sorted(headers)), 400
		if not required_headers.issubset(headers):
			missing = ", ".join(sorted(required_headers - headers))
			return render_template("admin/event_form.html", event=config, is_new=False,
								   error=f"The file is missing required column(s): {missing}.",
								   has_template=template_present, has_csv=has_csv,
								   csv_columns=sorted(headers)), 400
	except Exception:
		return render_template("admin/event_form.html", event=config, is_new=False,
							   error="Could not read that participant list.",
							   has_template=template_present, has_csv=has_csv,
							   csv_columns=csv_headers(slug)), 400

	try:
		save_event_csv(slug, content, source=(raw_upload, ext))
	except Exception:
		return render_template("admin/event_form.html", event=config, is_new=False,
							   error="Could not save the participant list to storage. Check the Supabase settings and try again.",
							   has_template=template_present, has_csv=has_csv,
							   csv_columns=csv_headers(slug)), 502

	converted = " Converted from the first sheet of the workbook." if ext == ".xlsx" else ""
	return render_template("admin/event_form.html", event=config, is_new=False,
						   success=f"Participant list uploaded.{converted}", error=None,
						   has_template=template_present, has_csv=True,
						   csv_columns=csv_headers(slug))


@app.route("/admin/events/<slug>/toggle", methods=["POST"])
@require_admin
def admin_toggle_event(slug: str):
	if not safe_slug(slug):
		return redirect(url_for("admin_dashboard"))
	config = load_event(slug)
	if config is None:
		return redirect(url_for("admin_dashboard"))
	config["active"] = not config.get("active", False)
	save_event_config(slug, config)
	_set_event_state(slug, active=config["active"], deleted=False)
	return redirect(url_for("admin_dashboard"))


@app.route("/admin/events/<slug>/send_emails", methods=["GET", "POST"])
@require_admin
def admin_send_emails(slug: str):
	if not _event_exists(slug):
		return redirect(url_for("admin_dashboard"))

	event_config = load_event(slug)

	if request.method == "GET":
		return render_template(
			"admin/email_form.html",
			event=event_config
		)

	subject_template = request.form.get("subject_template")
	plain_body_template = request.form.get("plain_body_template")
	html_body_template = request.form.get("html_body_template")

	# Update coordinates if they were changed
	if request.form.get("text_x"):
		event_config["text_x"] = int(request.form.get("text_x"))
	if request.form.get("text_y"):
		event_config["text_y"] = int(request.form.get("text_y"))
	if request.form.get("font_size"):
		event_config["font_size"] = int(request.form.get("font_size"))
	if request.form.get("font_color"):
		color_hex = request.form.get("font_color").lstrip("#")
		event_config["font_color"] = [int(color_hex[0:2], 16), int(color_hex[2:4], 16), int(color_hex[4:6], 16)]
	if request.form.get("font_key"):
		event_config["font_key"] = normalize_font_key(request.form.get("font_key"))

	save_event_config(slug, event_config)

	import threading
	try:
		from manage import bulk_generate

		def background_task():
			try:
				bulk_generate(slug, send_emails=True,
							  subject_template=subject_template,
							  plain_body_template=plain_body_template,
							  html_body_template=html_body_template)
			except Exception as e:
				print(f"Background email task failed: {e}")

		thread = threading.Thread(target=background_task)
		thread.daemon = True
		thread.start()

		return "<script>alert('Email sending has started in the background! Check server logs for progress.'); window.location.href='/admin/logs';</script>"
	except ImportError:
		return "manage.py module not found", 500


@app.route("/admin/logs", methods=["GET"])
@require_admin
def admin_logs():
	"""View the system.log file generated by background tasks."""
	logs_content = ""
	if os.path.exists(LOG_FILE):
		try:
			# Only the tail is useful, and the file is never rotated.
			with open(LOG_FILE, "r", encoding="utf-8", errors="replace") as f:
				logs_content = "".join(f.readlines()[-2000:])
		except Exception as e:
			logs_content = f"Error reading logs: {e}"
	return render_template("admin/logs.html", logs=logs_content)

@app.route("/admin/logs/clear", methods=["POST"])
@require_admin
def admin_clear_logs():
	"""Clear the system.log file."""
	if os.path.exists(LOG_FILE):
		try:
			with open(LOG_FILE, "w", encoding="utf-8") as f:
				f.write("")
		except Exception as e:
			print(f"Error clearing logs: {e}")
	return redirect(url_for("admin_logs"))


@app.route("/admin/events/<slug>/delete", methods=["POST"])
@require_admin
def admin_delete_event(slug: str):
	if not safe_slug(slug):
		return redirect(url_for("admin_dashboard"))
	if request.form.get("confirm", "") != slug:
		return redirect(url_for("admin_dashboard"))
	delete_event_storage(slug)
	_set_event_state(slug, deleted=True, active=False)
	return redirect(url_for("admin_dashboard"))


@app.route("/admin/events/<slug>/coordinates", methods=["GET"])
@require_admin
def admin_coordinate_editor(slug: str):
	"""Full-screen field placement editor. Draws and saves the event's fields list."""
	if not safe_slug(slug):
		return redirect(url_for("admin_dashboard"))
	config = load_event(slug)
	if config is None:
		return redirect(url_for("admin_dashboard"))
	dataset = load_participant_dataset(slug)
	# A handful of real rows drive the editor's live preview and the sample-picker,
	# so a club sees its actual data laid out rather than lorem. Admin-only, and the
	# admin can already see the whole CSV, so this leaks nothing new.
	sample_rows = dataset.raw_rows[:25]
	fonts = [
		{"key": o["key"], "label": o["label"], "family": o["css_family"],
		 "weight": o["css_weight"], "url": url_for("font_asset", font_key=o["key"])}
		for o in available_font_options()
	]
	return render_template(
		"admin/coordinate_editor.html",
		event=config,
		fields_json=_json_for_script(normalize_fields(config)),
		columns_json=_json_for_script(dataset.columns),
		sample_rows_json=_json_for_script(sample_rows),
		fonts_json=_json_for_script(fonts),
		max_fields=MAX_FIELDS,
		has_csv=_event_csv_exists(slug),
		save_url=url_for("admin_save_fields", slug=slug),
		template_url=url_for("admin_template_preview", slug=slug),
		back_url=url_for("admin_edit_event", slug=slug),
	)


@app.route("/admin/events/<slug>/fields", methods=["POST"])
@require_admin
def admin_save_fields(slug: str):
	"""Persist the editor's field list. JSON in, JSON out, CSRF via header."""
	if not safe_slug(slug):
		return jsonify({"ok": False, "error": "Unknown event."}), 404
	config = load_event(slug)
	if config is None:
		return jsonify({"ok": False, "error": "Unknown event."}), 404
	payload = request.get_json(silent=True) or {}
	fields, error = validate_fields_payload(slug, payload.get("fields"))
	if error:
		return jsonify({"ok": False, "error": error}), 400
	config["fields"] = fields
	save_event_config(slug, config)
	return jsonify({"ok": True, "message": "Placement saved.", "fields": fields})


@app.route("/admin/events/<slug>/template-preview", methods=["GET"])
@require_admin
def admin_template_preview(slug: str):
	"""Serve certificate template image for canvas preview in event editor."""
	if not safe_slug(slug):
		return "Not found", 404
	config = load_event(slug)
	if config is None:
		return "Not found", 404
	data = load_template_bytes(slug, config)
	if data is None:
		return "Template not found", 404
	ext = template_ext_for(slug, config)
	response = send_file(
		BytesIO(data),
		mimetype=TEMPLATE_CONTENT_TYPES.get(ext, "application/octet-stream"),
		etag=template_version_for(slug, config),
	)
	response.cache_control.private = True
	response.cache_control.max_age = 60
	return response


@app.route("/admin/events/<slug>/render-preview", methods=["GET"])
@require_admin
def admin_render_preview(slug: str):
	"""Render a preview with the same server-side Pillow path used for generated certificates."""
	if not safe_slug(slug):
		return "Not found", 404
	config = load_event(slug)
	if config is None:
		return "Not found", 404
	try:
		with render_slot(slug):
			image = get_template_image(slug, config)
			if image is None:
				return "Template not found", 404
			metadata = build_preview_metadata(config, request.args.get("cert_name"))
			draw_name_on_image(image, metadata)
			image_bytes, mimetype = encode_certificate(image, variant="preview")
	except RenderCapacityError:
		# Admin previews are the same CPU as participant renders, so they queue in
		# the same pool rather than competing with it from outside.
		return render_busy_response()
	response = send_file(BytesIO(image_bytes), mimetype=mimetype)
	# Short cache for admin previews (they may change as settings are edited)
	response.cache_control.max_age = 60
	response.cache_control.public = True
	return response


@app.route("/healthz", methods=["GET"])
def healthz():
	"""
	Keep-alive endpoint for the warm-up cron.

	It deliberately touches Supabase: free Supabase projects pause after about a
	week of inactivity, and the app only reads storage on a cache miss, so pinging
	Flask alone would not be enough to keep the storage project awake.
	"""
	storage = "disabled"
	if _supabase_enabled():
		try:
			_supabase_ping()
			storage = "ok"
		except Exception:
			storage = "error"
	return jsonify({"ok": True, "storage": storage})


@app.route("/assets/fonts/<font_key>.ttf", methods=["GET"])
def font_asset(font_key: str):
	"""Serve event font files used by PIL so browser previews match generated files."""
	if normalize_font_key(font_key) != font_key:
		return "Font not found", 404
	font_option = resolve_font_option(font_key)
	if not os.path.exists(font_option["path"]):
		return "Font not found", 404
	response = send_file(font_option["path"], mimetype="font/ttf")
	# Cache fonts for 1 year (rarely change)
	response.cache_control.max_age = 31536000
	response.cache_control.public = True
	response.cache_control.immutable = True
	return response


@app.route("/assets/fonts/montserrat-bold.ttf", methods=["GET"])
def montserrat_bold_font():
	"""Backward-compatible route for existing CSS references."""
	return font_asset(DEFAULT_FONT_KEY)


# ─── Entry point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
	app.run(debug=True)
